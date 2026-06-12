from __future__ import annotations

import csv
from datetime import date
from io import BytesIO, StringIO
import json
import os
from pathlib import Path
import tempfile
from threading import Event
import time
import unittest
from unittest.mock import patch

import pandas as pd
from fastapi.testclient import TestClient

from mpstats_app.config import AppSettings
from mpstats_app.main import create_app
from mpstats_app.repositories.duckdb_repository import DuckDbAppRepository
from mpstats_app.services.smart_pipeline_service import month_day_coverage
from mpstats_app.utils import quote_duckdb_name
from pipeline.repositories.file_repository import write_semicolon_csv
from pipeline.repositories.sql_repository import connect


def make_settings(root: Path) -> AppSettings:
    return AppSettings.create(
        project_root=root,
        workdir=root / "pipeline",
        db_path=root / "mpstats.duckdb",
        config_path=root / "pipeline" / "step1_export_config.json",
        rules_path=root / "classifiers" / "rules.csv",
        scheduler_poll_seconds=0.1,
        static_dir=root / "web" / "dist",
    )


def seed_project(root: Path) -> None:
    (root / "pipeline").mkdir(parents=True)
    (root / "classifiers").mkdir(parents=True)
    (root / "pipeline" / "step1_export_config.json").write_text(
        """
        {
          "export_months_by_year": {"2025": [1]},
          "save_dir": "",
          "skip_if_exists": true,
          "extract_zip": true,
          "cookie": "",
          "tasks": [
            {"active": true, "mp": "oz", "path": "Продукты/Тест", "cat": "Тест", "fbs": 1}
          ]
        }
        """,
        encoding="utf-8",
    )
    (root / "classifiers" / "rules.csv").write_text(
        "\n".join(
            [
                "active;priority;category;target_column;match_field;match_type;pattern;set_value;mode;comment;conditions_json",
                "1;1;*;Тип;Название;contains;лимон;Кислота;fill_empty;;",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    (root / "Справочник категорий MP STATS.csv").write_text(
        "\n".join(
            [
                "Чек;Категория;МП;FBS;От;До;Комментарий;Путь;Фильтр;Путь2;Фильтр2;Актуализация",
                ";Лимонная кислота;Озон;1;2025;2025;;Продукты/Тест/Лимонная кислота;;;;",
                ";Лимонная кислота;WB;;2025;2025;;Продукты/Тест/WB;;;;",
                ";Масло;WB;;2025;2025;;Продукты/Тест/Масло;\"\"\"Подсолнечное\"\"\";;;",
                ";Яндекс тест;ЯМ;1;2025;2025;;Продукты/Тест/Яндекс;;;;",
                ";Смена пути;WB;;янв.25;янв.25;;Продукты/Старый путь;;;;",
                ";Смена пути;WB;;фев.25;фев.25;;Продукты/Новый путь;;;;",
                ";Пустой путь;WB;;2025;2025;;НД;;;;",
            ]
        )
        + "\n",
        encoding="utf-8-sig",
    )


def _has_openpyxl() -> bool:
    try:
        __import__("openpyxl")
        return True
    except ModuleNotFoundError:
        return False


class WebApiTest(unittest.TestCase):
    def test_month_day_coverage_marks_current_month_by_saved_day(self) -> None:
        current = month_day_coverage(2026, 5, today=date(2026, 5, 28))
        self.assertEqual(current["days_loaded"], 28)
        self.assertEqual(current["days_in_month"], 31)
        self.assertEqual(current["data_actual_until"], "2026-05-28")

        past = month_day_coverage(2026, 4, today=date(2026, 5, 28))
        self.assertEqual(past["days_loaded"], 30)
        self.assertEqual(past["days_in_month"], 30)
        self.assertEqual(past["data_actual_until"], "2026-04-30")

    def test_weight_column_migration_renames_unit_and_total_weight(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            with connect(settings.db_path) as con:
                con.execute(
                    f"""
                    CREATE TABLE {settings.products_table} (
                        SKU VARCHAR,
                        {quote_duckdb_name("Вес, кг")} DOUBLE,
                        {quote_duckdb_name("Вес, кг (сумм.)")} DOUBLE
                    )
                    """
                )
                con.execute(f"INSERT INTO {settings.products_table} VALUES ('sku-pack', 0.5, 1.5)")

            DuckDbAppRepository(settings).ensure_ready()

            with connect(settings.db_path) as con:
                columns = [row[1] for row in con.execute(f"PRAGMA table_info({settings.products_table})").fetchall()]
                row = con.execute(
                    f"""
                    SELECT
                        TRY_CAST(REPLACE(CAST({quote_duckdb_name("Вес, кг")} AS VARCHAR), ',', '.') AS DOUBLE),
                        TRY_CAST(REPLACE(CAST({quote_duckdb_name("Вес, кг (ед.)")} AS VARCHAR), ',', '.') AS DOUBLE)
                    FROM {settings.products_table}
                    """
                ).fetchone()

            self.assertIn("Вес, кг", columns)
            self.assertIn("Вес, кг (ед.)", columns)
            self.assertNotIn("Вес, кг (сумм.)", columns)
            self.assertEqual(row, (1.5, 0.5))

    def test_db_import_maps_legacy_weight_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)
            repository.ensure_ready()

            legacy_file = root / "legacy-weight.csv"
            write_semicolon_csv(
                pd.DataFrame([{"SKU": "sku-pack", "Вес, кг": "0,5", "Вес, кг (сумм.)": "1,5", "Продажи, шт": 1}]),
                legacy_file,
            )
            repository.import_products_file_idempotent(
                run_id="run-legacy-weight",
                csv_path=legacy_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=2,
                marketplace_code="oz",
                category_key="sugar",
            )

            with connect(settings.db_path) as con:
                columns = [row[1] for row in con.execute(f"PRAGMA table_info({settings.products_table})").fetchall()]
                row = con.execute(
                    f"""
                    SELECT
                        TRY_CAST(REPLACE(CAST({quote_duckdb_name("Вес, кг")} AS VARCHAR), ',', '.') AS DOUBLE),
                        TRY_CAST(REPLACE(CAST({quote_duckdb_name("Вес, кг (ед.)")} AS VARCHAR), ',', '.') AS DOUBLE)
                    FROM {settings.products_table}
                    """
                ).fetchone()

            self.assertNotIn("Вес, кг (сумм.)", columns)
            self.assertEqual(row, (1.5, 0.5))

    def test_db_import_widens_empty_classifier_column_to_text(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)

            first_file = root / "first.csv"
            write_semicolon_csv(
                pd.DataFrame([{"SKU": "sku-1", "Подкатегория": None, "Цена за кг": 10.5}]),
                first_file,
            )
            repository.import_products_file_idempotent(
                run_id="run-1",
                csv_path=first_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=2,
                marketplace_code="oz",
                category_key="sugar",
            )

            second_file = root / "second.csv"
            write_semicolon_csv(
                pd.DataFrame([{"SKU": "sku-2", "Подкатегория": "Тростниковый", "Цена за кг": 12.5}]),
                second_file,
            )
            inserted = repository.import_products_file_idempotent(
                run_id="run-1",
                csv_path=second_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="sugar",
            )

            self.assertEqual(inserted, 1)
            with connect(settings.db_path) as con:
                data_type = con.execute(
                    """
                    SELECT data_type
                    FROM information_schema.columns
                    WHERE table_name = ? AND column_name = 'Подкатегория'
                    """,
                    [settings.products_table],
                ).fetchone()[0]
                values = con.execute(
                    f'SELECT "Подкатегория" FROM "{settings.products_table}" ORDER BY "SKU"'
                ).fetchall()

            self.assertEqual(data_type, "VARCHAR")
            self.assertEqual(values, [(None,), ("Тростниковый",)])

    def test_db_import_preserves_date_column(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)

            source_file = root / "products.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {
                            "Дата": "01.01.2025",
                            "SKU": "sku-1",
                            "Продажи, шт": "5",
                            "Объем, кг": "1,5",
                        }
                    ]
                ),
                source_file,
            )

            inserted = repository.import_products_file_idempotent(
                run_id="run-1",
                csv_path=source_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="sugar",
            )

            self.assertEqual(inserted, 1)
            with connect(settings.db_path) as con:
                values = con.execute(f'SELECT "Дата" FROM "{settings.products_table}"').fetchall()
            self.assertEqual(values, [("01.01.2025",)])

    def test_db_import_filters_zero_sales_and_volume_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)

            source_file = root / "products.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {"SKU": "valid", "Продажи, шт": "5", "Объем, т": "0,25"},
                        {"SKU": "zero-sales", "Продажи, шт": "0", "Объем, т": "0,25"},
                        {"SKU": "zero-volume", "Продажи, шт": "5", "Объем, т": "0"},
                        {"SKU": "bad-sales", "Продажи, шт": "мусор", "Объем, т": "0,25"},
                        {"SKU": "bad-volume", "Продажи, шт": "5", "Объем, т": "мусор"},
                    ]
                ),
                source_file,
            )

            inserted = repository.import_products_file_idempotent(
                run_id="run-1",
                csv_path=source_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="sugar",
            )

            self.assertEqual(inserted, 1)
            with connect(settings.db_path) as con:
                rows = con.execute(f"SELECT SKU FROM {settings.products_table}").fetchall()
                null_metadata = con.execute(
                    f"""
                    SELECT COUNT(*)
                    FROM {settings.products_table}
                    WHERE {quote_duckdb_name('__project_name')} IS NULL
                       OR {quote_duckdb_name('__year')} IS NULL
                       OR {quote_duckdb_name('__month')} IS NULL
                       OR {quote_duckdb_name('__marketplace_code')} IS NULL
                       OR {quote_duckdb_name('__category_key')} IS NULL
                       OR {quote_duckdb_name('__row_hash')} IS NULL
                    """
                ).fetchone()[0]
            self.assertEqual(rows, [("valid",)])
            self.assertEqual(null_metadata, 0)

    def test_db_import_idempotent_rerun_does_not_duplicate_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)

            source_file = root / "products.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {"SKU": "sku-1", "Продажи, шт": "5", "Объем, кг": "1.5"},
                        {"SKU": "sku-2", "Продажи, шт": "7", "Объем, кг": "2.5"},
                    ]
                ),
                source_file,
            )

            first_inserted = repository.import_products_file_idempotent(
                run_id="run-1",
                csv_path=source_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="sugar",
            )
            second_inserted = repository.import_products_file_idempotent(
                run_id="run-2",
                csv_path=source_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="sugar",
            )

            self.assertEqual(first_inserted, 2)
            self.assertEqual(second_inserted, 2)
            with connect(settings.db_path) as con:
                total = con.execute(f"SELECT COUNT(*) FROM {settings.products_table}").fetchone()[0]
                duplicate_hashes = con.execute(
                    f"""
                    SELECT COUNT(*)
                    FROM (
                        SELECT {quote_duckdb_name('__row_hash')}, COUNT(*) AS cnt
                        FROM {settings.products_table}
                        GROUP BY {quote_duckdb_name('__row_hash')}
                        HAVING COUNT(*) > 1
                    )
                    """
                ).fetchone()[0]
            self.assertEqual(total, 2)
            self.assertEqual(duplicate_hashes, 0)

    def test_db_import_blocks_duplicate_month_category_marketplace_slice(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)

            first_file = root / "first.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {
                            "Маркетплейс": "Ozon",
                            "Категория": "Сахар",
                            "SKU": "sku-1",
                            "Продажи, шт": "5",
                            "Объем, кг": "1.5",
                        }
                    ]
                ),
                first_file,
            )
            first_inserted = repository.import_products_file_idempotent(
                run_id="run-1",
                csv_path=first_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="sugar-old",
                category_name="Сахар",
            )
            repository.upsert_cube_entry(
                {
                    "project_name": "unit",
                    "year": 2025,
                    "month": 1,
                    "marketplace": "Ozon",
                    "marketplace_code": "oz",
                    "category_key": "sugar-old",
                    "category_name": "Сахар",
                    "rows_count": first_inserted,
                    "source_processed_file_path": str(first_file),
                    "file_hash": "first",
                }
            )

            second_file = root / "second.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {
                            "Маркетплейс": "Ozon",
                            "Категория": "Сахар",
                            "SKU": "sku-2",
                            "Продажи, шт": "7",
                            "Объем, кг": "2.5",
                        }
                    ]
                ),
                second_file,
            )
            with self.assertRaisesRegex(ValueError, "Срез уже сохранён"):
                repository.import_products_file_idempotent(
                    run_id="run-2",
                    csv_path=second_file,
                    table_name=settings.products_table,
                    project_name="unit",
                    year=2025,
                    month=1,
                    marketplace_code="oz",
                    category_key="sugar-new",
                    category_name="Сахар",
                )

            with connect(settings.db_path) as con:
                total_after_block = con.execute(f"SELECT COUNT(*) FROM {settings.products_table}").fetchone()[0]
                loads_after_block = con.execute("SELECT COUNT(*) FROM pipeline_loads").fetchone()[0]
            self.assertEqual(total_after_block, 1)
            self.assertEqual(loads_after_block, 1)

            overwrite_inserted = repository.import_products_file_idempotent(
                run_id="run-3",
                csv_path=second_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="sugar-new",
                category_name="Сахар",
                overwrite=True,
            )
            repository.upsert_cube_entry(
                {
                    "project_name": "unit",
                    "year": 2025,
                    "month": 1,
                    "marketplace": "Ozon",
                    "marketplace_code": "oz",
                    "category_key": "sugar-new",
                    "category_name": "Сахар",
                    "rows_count": overwrite_inserted,
                    "source_processed_file_path": str(second_file),
                    "file_hash": "second",
                }
            )

            self.assertEqual(overwrite_inserted, 1)
            with connect(settings.db_path) as con:
                rows = con.execute(f"SELECT SKU FROM {settings.products_table}").fetchall()
                registry_total = con.execute("SELECT COUNT(*) FROM cube_registry").fetchone()[0]
                registry_key = con.execute("SELECT category_key FROM cube_registry").fetchone()[0]
            self.assertEqual(rows, [("sku-2",)])
            self.assertEqual(registry_total, 1)
            self.assertEqual(registry_key, "sugar-new")

    def test_db_import_deduplicates_business_rows_across_source_types(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)

            category_file = root / "category.csv"
            subject_file = root / "subject.csv"
            category_row = pd.DataFrame(
                [
                    {
                        "Маркетплейс": "WB",
                        "SKU": "sku-1",
                        "Название": "лимон дубль",
                        "Бренд": "Brand",
                        "Продажи, шт": "5",
                        "Выручка, руб": "50",
                    }
                ]
            )
            subject_row = category_row.copy()
            write_semicolon_csv(category_row, category_file)
            write_semicolon_csv(subject_row, subject_file)

            first_inserted = repository.import_products_file_idempotent(
                run_id="run-category",
                csv_path=category_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="wb",
                category_key="category-key",
                source_type="category",
            )
            with connect(settings.db_path) as con:
                con.execute(f"UPDATE {settings.products_table} SET {quote_duckdb_name('__business_row_hash')} = NULL")
            second_inserted = repository.import_products_file_idempotent(
                run_id="run-subject",
                csv_path=subject_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="wb",
                category_key="subject-key",
                source_type="subject",
            )

            self.assertEqual(first_inserted, 1)
            self.assertEqual(second_inserted, 0)
            with connect(settings.db_path) as con:
                total = con.execute(f"SELECT COUNT(*) FROM {settings.products_table}").fetchone()[0]
                source_types = con.execute(
                    f"SELECT {quote_duckdb_name('__source_type')}, COUNT(*) FROM {settings.products_table} GROUP BY 1"
                ).fetchall()
                business_hashes = con.execute(
                    f"SELECT COUNT(DISTINCT {quote_duckdb_name('__business_row_hash')}) FROM {settings.products_table}"
                ).fetchone()[0]

            self.assertEqual(total, 1)
            self.assertEqual(source_types, [("category", 1)])
            self.assertEqual(business_hashes, 1)

    def test_db_import_rolls_back_slice_replace_on_insert_error(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)

            source_file = root / "products.csv"
            write_semicolon_csv(
                pd.DataFrame([{"SKU": "sku-1", "Продажи, шт": "5", "Объем, кг": "1.5"}]),
                source_file,
            )
            inserted = repository.import_products_file_idempotent(
                run_id="run-1",
                csv_path=source_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="sugar",
            )
            self.assertEqual(inserted, 1)

            replacement_file = root / "replacement.csv"
            write_semicolon_csv(
                pd.DataFrame([{"SKU": "sku-2", "Продажи, шт": "9", "Объем, кг": "3.5"}]),
                replacement_file,
            )
            with patch(
                "mpstats_app.repositories.duckdb_repository._insert_stage_sql",
                return_value="INSERT INTO missing_table SELECT 1",
            ):
                with self.assertRaises(Exception):
                    repository.import_products_file_idempotent(
                        run_id="run-2",
                        csv_path=replacement_file,
                        table_name=settings.products_table,
                        project_name="unit",
                        year=2025,
                        month=1,
                        marketplace_code="oz",
                        category_key="sugar",
                    )

            with connect(settings.db_path) as con:
                rows = con.execute(f"SELECT SKU FROM {settings.products_table}").fetchall()
                loads = con.execute("SELECT COUNT(*) FROM pipeline_loads").fetchone()[0]
            self.assertEqual(rows, [("sku-1",)])
            self.assertEqual(loads, 1)

    def test_health_settings_schedules_and_run_queue(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            app = create_app(settings, start_workers=False)

            with TestClient(app) as client:
                health = client.get("/api/health")
                self.assertEqual(health.status_code, 200)
                self.assertTrue(health.json()["ok"])

                config_response = client.get("/api/settings/export-config")
                self.assertEqual(config_response.status_code, 200)
                self.assertEqual(config_response.json()["config"]["export_months_by_year"], {"2025": [1]})

                rules_response = client.put("/api/rules", json={"content": "active;priority\n"})
                self.assertEqual(rules_response.status_code, 200)
                self.assertEqual((root / "classifiers" / "rules.csv").read_text(encoding="utf-8"), "active;priority\n")

                structured_rules = client.put(
                    "/api/classifier/rules",
                    json={
                        "rules": [
                            {
                                "active": True,
                                "priority": 10,
                                "category": "*",
                                "target_column": "Тип",
                                "set_value": "Кислота",
                                "mode": "fill_empty",
                                "comment": "unit",
                                "conditions": [
                                    {"join_with_prev": "and", "match_field": "Название", "match_type": "contains", "pattern": "лимон"},
                                    {"join_with_prev": "or", "match_field": "SKU", "match_type": "contains", "pattern": "лимон"},
                                ],
                            },
                            {
                                "active": True,
                                "priority": 999,
                                "category": "*",
                                "target_column": "Тип",
                                "set_value": "Прочее",
                                "mode": "fill_empty",
                                "comment": "fallback",
                                "conditions": [
                                    {"join_with_prev": "and", "match_field": "", "match_type": "otherwise", "pattern": ""},
                                ],
                            }
                        ]
                    },
                )
                self.assertEqual(structured_rules.status_code, 200)
                rules_payload = structured_rules.json()["rules"]
                self.assertEqual(rules_payload[0]["conditions"][1]["join_with_prev"], "or")
                self.assertEqual(rules_payload[1]["conditions"][0]["match_type"], "otherwise")
                self.assertIn("conditions_json", (root / "classifiers" / "rules.csv").read_text(encoding="utf-8-sig"))

                manual_response = client.put(
                    "/api/classifier/manual-overrides",
                    json={
                        "overrides": [
                            {
                                "active": True,
                                "priority": 10,
                                "match_field": "Артикул",
                                "match_value": "123",
                                "target_column": "Подкатегория",
                                "set_value": "Ручная",
                                "mode": "overwrite",
                                "comment": "unit",
                            }
                        ]
                    },
                )
                self.assertEqual(manual_response.status_code, 200)
                manual_payload = manual_response.json()["overrides"]
                self.assertEqual(manual_payload[0]["match_field"], "Артикул")
                self.assertEqual(manual_payload[0]["set_value"], "Ручная")
                self.assertIn("manual_overrides.csv", manual_response.json()["path"])
                self.assertIn("Артикул;123;Подкатегория;Ручная", (root / "classifiers" / "manual_overrides.csv").read_text(encoding="utf-8-sig"))

                schedule_response = client.post(
                    "/api/schedules",
                    json={
                        "name": "daily",
                        "project_name": "unit",
                        "steps": "2-6",
                        "enabled": True,
                        "interval_minutes": 60,
                    },
                )
                self.assertEqual(schedule_response.status_code, 200)
                schedule_id = schedule_response.json()["schedule_id"]
                self.assertEqual(client.get("/api/schedules").json()["schedules"][0]["schedule_id"], schedule_id)

                run_response = client.post("/api/runs", json={"project_name": "unit", "steps": "2-3"})
                self.assertEqual(run_response.status_code, 200)
                self.assertEqual(run_response.json()["status"], "queued")

    def test_project_management_lists_and_deletes_project(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            app = create_app(settings, start_workers=False)

            with TestClient(app) as client:
                created = client.post("/api/projects", json={"project_name": "Новый проект"})
                self.assertEqual(created.status_code, 200)
                self.assertEqual(created.json()["project_name"], "Новый проект")
                self.assertTrue((root / "data" / "projects" / "Новый_проект").exists())
                self.assertFalse(created.json()["has_files"])

                after_create = client.get("/api/projects")
                self.assertIn("Новый проект", {project["project_name"] for project in after_create.json()["projects"]})
                self.assertTrue(
                    next(project for project in after_create.json()["projects"] if project["project_name"] == "Новый проект")["is_current"]
                )

                repository: DuckDbAppRepository = app.state.repository
                repository.create_pipeline_run(
                    run_id="plan-1",
                    project_name="unit",
                    run_type="historical_backfill",
                    period_from="2025-01",
                    period_to="2025-01",
                    selected_category_ids=["category-1"],
                    settings={},
                )
                csv_path = root / "classified.csv"
                write_semicolon_csv(
                    pd.DataFrame([{"Маркетплейс": "Ozon", "Категория": "Тест", "Название": "Товар"}]),
                    csv_path,
                )
                inserted = repository.import_products_file_idempotent(
                    run_id="plan-1",
                    csv_path=csv_path,
                    table_name=settings.products_table,
                    project_name="unit",
                    year=2025,
                    month=1,
                    marketplace_code="oz",
                    category_key="test",
                )
                repository.upsert_cube_entry(
                    {
                        "project_name": "unit",
                        "year": 2025,
                        "month": 1,
                        "marketplace": "Ozon",
                        "marketplace_code": "oz",
                        "category_key": "test",
                        "category_name": "Тест",
                        "rows_count": inserted,
                        "days_loaded": 17,
                        "days_in_month": 31,
                        "data_actual_until": "2025-01-17",
                        "source_processed_file_path": "unit.csv",
                        "file_hash": "hash",
                    }
                )
                project_dir = root / "data" / "projects" / "unit"
                (project_dir / "raw").mkdir(parents=True)
                (project_dir / "raw" / "unit.csv").write_text("x", encoding="utf-8")

                projects = client.get("/api/projects")
                self.assertEqual(projects.status_code, 200)
                unit = next(project for project in projects.json()["projects"] if project["project_name"] == "unit")
                self.assertEqual(unit["cube_slices_count"], 1)
                self.assertEqual(unit["product_rows_count"], 1)
                self.assertTrue(unit["has_files"])
                cube = client.get("/api/workflow/pipeline/cube", params={"project_name": "unit"})
                self.assertEqual(cube.status_code, 200)
                self.assertEqual(cube.json()["total"], 1)
                cube_item = cube.json()["items"][0]
                self.assertEqual(cube_item["days_loaded"], 17)
                self.assertEqual(cube_item["days_in_month"], 31)
                self.assertEqual(cube_item["data_actual_until"], "2025-01-17")
                self.assertIsNotNone(cube_item["exported_at"])

                deleted = client.delete("/api/projects", params={"project_name": "unit", "delete_files": True})
                self.assertEqual(deleted.status_code, 200)
                self.assertEqual(deleted.json()["deleted"]["cube_registry"], 1)
                self.assertEqual(deleted.json()["deleted"]["product_rows"], 1)
                self.assertFalse(project_dir.exists())

                after_delete = client.get("/api/projects")
                self.assertNotIn("unit", {project["project_name"] for project in after_delete.json()["projects"]})

    def test_spa_fallback_does_not_shadow_unknown_api_paths(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            static_dir = root / "web" / "dist"
            static_dir.mkdir(parents=True)
            (static_dir / "index.html").write_text("<div id=\"root\"></div>", encoding="utf-8")
            settings = make_settings(root)
            app = create_app(settings, start_workers=False)

            with TestClient(app) as client:
                api_response = client.get("/api/workflow/pipeline/missing")
                self.assertEqual(api_response.status_code, 404)
                self.assertEqual(api_response.headers["content-type"].split(";")[0], "application/json")
                self.assertEqual(api_response.json()["detail"], "API endpoint not found")

                page_response = client.get("/workflow/monthly")
                self.assertEqual(page_response.status_code, 200)
                self.assertIn("root", page_response.text)
                self.assertEqual(
                    page_response.headers["cache-control"],
                    "no-store, no-cache, must-revalidate, max-age=0",
                )
                self.assertEqual(page_response.headers["pragma"], "no-cache")
                self.assertEqual(page_response.headers["expires"], "0")

    def test_product_search_uses_latest_successful_run(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)
            repository.ensure_ready()
            run = repository.create_run(
                run_id="run-1",
                project_name="unit",
                steps="2,3,4,5,6",
                source="manual",
                schedule_id=None,
                workdir=settings.workdir,
                config_path=settings.config_path,
                rules_path=settings.rules_path,
                db_path=settings.db_path,
                products_table=settings.products_table,
                write_xlsx=False,
                max_weight_kg=40.0,
                fill_unclassified=None,
            )
            repository.finish_run(str(run["run_id"]), "succeeded")
            csv_path = root / "classified.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {
                            "Маркетплейс": "Ozon",
                            "Категория": "Тест",
                            "Артикул": "123",
                            "SKU": "Лимонная кислота 1 кг",
                            "Бренд": "Brand",
                        }
                    ]
                ),
                csv_path,
            )
            repository.import_products_file(
                run_id="run-1",
                csv_path=csv_path,
                table_name=settings.products_table,
                project_name="unit",
            )
            other_run = repository.create_run(
                run_id="run-2",
                project_name="other",
                steps="2,3,4,5,6",
                source="manual",
                schedule_id=None,
                workdir=settings.workdir,
                config_path=settings.config_path,
                rules_path=settings.rules_path,
                db_path=settings.db_path,
                products_table=settings.products_table,
                write_xlsx=False,
                max_weight_kg=40.0,
                fill_unclassified=None,
            )
            repository.finish_run(str(other_run["run_id"]), "succeeded")
            other_csv_path = root / "other_classified.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {
                            "Маркетплейс": "Ozon",
                            "Категория": "Тест",
                            "Артикул": "999",
                            "SKU": "Соль 1 кг",
                            "Бренд": "Other",
                        }
                    ]
                ),
                other_csv_path,
            )
            repository.import_products_file(
                run_id="run-2",
                csv_path=other_csv_path,
                table_name=settings.products_table,
                project_name="other",
            )

            app = create_app(settings, start_workers=False)
            with TestClient(app) as client:
                preview = client.get("/api/products", params={"project_name": "unit", "limit": 100})
                self.assertEqual(preview.status_code, 200)
                preview_payload = preview.json()
                self.assertEqual(preview_payload["total"], 1)
                self.assertEqual(preview_payload["run_id"], "run-1")
                self.assertEqual(str(preview_payload["rows"][0]["Артикул"]), "123")

                response = client.get("/api/products", params={"project_name": "unit", "query": "лимон"})
                self.assertEqual(response.status_code, 200)
                payload = response.json()
                self.assertEqual(payload["total"], 1)
                self.assertEqual(payload["run_id"], "run-1")
                self.assertEqual(str(payload["rows"][0]["Артикул"]), "123")

    def test_export_options_preview_build_split_and_by_category(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)
            repository.ensure_ready()

            lemon_file = root / "lemon.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {"Маркетплейс": "Ozon", "Категория": "Лимонная кислота", "SKU": "sku-1", "Название": "лимон 1 кг", "Бренд": "Brand A", "Продажи, шт": 3, "Вес, кг": "1,0", "Вес, кг (ед.)": "1,0"},
                        {"Маркетплейс": "Ozon", "Категория": "Лимонная кислота", "SKU": "sku-2", "Название": "лимон 2 кг", "Бренд": "Brand B", "Продажи, шт": 5, "Вес, кг": "1,5", "Вес, кг (ед.)": "0,5"},
                        {"Маркетплейс": "Ozon", "Категория": "Лимонная кислота", "SKU": "sku-0", "Название": "лимон без продаж", "Бренд": "Brand Z", "Продажи, шт": 0},
                    ]
                ),
                lemon_file,
            )
            repository.import_products_file_idempotent(
                run_id="run-export-1",
                csv_path=lemon_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="lemon-oz",
            )
            soap_file = root / "soap.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {"Маркетплейс": "WB", "Категория": "Мыло", "SKU": "sku-3", "Название": "мыло тест", "Бренд": "Brand C"},
                    ]
                ),
                soap_file,
            )
            repository.import_products_file_idempotent(
                run_id="run-export-2",
                csv_path=soap_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=2,
                marketplace_code="wb",
                category_key="soap-wb",
            )

            app = create_app(settings, start_workers=False)
            with (
                patch.object(app.state.repository, "count_export_products", side_effect=AssertionError("Export flow should reuse breakdown totals.")),
                TestClient(app) as client,
            ):
                options = client.get("/api/exports/options", params={"project_name": "unit"})
                self.assertEqual(options.status_code, 200)
                options_payload = options.json()
                self.assertEqual(options_payload["period_from"], "2025-01")
                self.assertEqual(options_payload["period_to"], "2025-02")
                self.assertIn("Название", options_payload["columns"])
                self.assertNotIn("__row_hash", options_payload["columns"])
                self.assertEqual({row["category_key"] for row in options_payload["categories"]}, {"lemon-oz", "soap-wb"})
                self.assertEqual(Path(options_payload["default_output_dir"]), (root / "data" / "projects" / "unit" / "exports").resolve())

                preview_payload = {
                    "project_name": "unit",
                    "category_keys": ["lemon-oz"],
                    "period_from": "2025-01",
                    "period_to": "2025-01",
                    "selected_columns": ["SKU", "Название", "Бренд", "Продажи, шт", "Вес, кг", "Вес, кг (ед.)"],
                    "filters": [
                        {"column": "Название", "match_type": "contains", "value": "лимон"},
                        {"column": "Продажи, шт", "match_type": "gt", "value": "0"},
                    ],
                    "excluded_row_hashes": [],
                    "sort_column": "SKU",
                    "sort_direction": "asc",
                    "split_by_category": False,
                    "limit": 100,
                    "offset": 0,
                }
                preview = client.post("/api/exports/preview", json=preview_payload)
                self.assertEqual(preview.status_code, 200)
                preview_data = preview.json()
                self.assertEqual(preview_data["total"], 2)
                self.assertEqual(preview_data["columns"], ["SKU", "Название", "Бренд", "Продажи, шт", "Вес, кг", "Вес, кг (ед.)"])
                self.assertEqual(preview_data["breakdown"][0]["period"], "2025-01")
                self.assertEqual(preview_data["breakdown"][0]["rows_count"], 2)
                excluded_hash = preview_data["rows"][0]["__row_hash"]

                excluded_preview = client.post(
                    "/api/exports/preview",
                    json={**preview_payload, "excluded_row_hashes": [excluded_hash]},
                )
                self.assertEqual(excluded_preview.status_code, 200)
                self.assertEqual(excluded_preview.json()["total"], 1)

                template = client.post(
                    "/api/exports/templates",
                    json={
                        **{key: value for key, value in preview_payload.items() if key not in {"limit", "offset", "excluded_row_hashes"}},
                        "name": "Ежемесячный лимон",
                        "export_format": "xlsx",
                        "output_dir": str(root / "template-exports"),
                    },
                )
                self.assertEqual(template.status_code, 200)
                template_payload = template.json()
                self.assertEqual(template_payload["name"], "Ежемесячный лимон")
                self.assertEqual(template_payload["project_name"], "unit")
                self.assertEqual(template_payload["category_keys"], ["lemon-oz"])
                self.assertEqual(template_payload["selected_columns"], ["SKU", "Название", "Бренд", "Продажи, шт", "Вес, кг", "Вес, кг (ед.)"])

                templates = client.get("/api/exports/templates", params={"project_name": "unit"})
                self.assertEqual(templates.status_code, 200)
                self.assertEqual([row["name"] for row in templates.json()["templates"]], ["Ежемесячный лимон"])

                deleted_template = client.delete(
                    f"/api/exports/templates/{template_payload['id']}",
                    params={"project_name": "unit"},
                )
                self.assertEqual(deleted_template.status_code, 200)
                self.assertEqual(client.get("/api/exports/templates", params={"project_name": "unit"}).json()["templates"], [])

                output_dir = root / "exports"
                with patch.object(
                    app.state.repository,
                    "fetch_export_products_dataframe",
                    side_effect=AssertionError("Raw XLSX export should use DuckDB COPY, not pandas/openpyxl batches."),
                ):
                    built = client.post(
                        "/api/exports/build",
                        json={
                            **preview_payload,
                            "excluded_row_hashes": [excluded_hash],
                            "output_dir": str(output_dir),
                            "confirm_large_export": False,
                        },
                    )
                self.assertEqual(built.status_code, 200)
                built_payload = built.json()
                self.assertEqual(built_payload["total"], 1)
                self.assertEqual(len(built_payload["artifacts"]), 1)
                self.assertEqual(Path(built_payload["output_dir"]), (output_dir / "unit").resolve())
                xlsx_path = Path(built_payload["artifacts"][0]["path"])
                self.assertEqual(xlsx_path.parent, (output_dir / "unit").resolve())
                self.assertTrue(xlsx_path.exists())

                from openpyxl import load_workbook

                workbook = load_workbook(xlsx_path, read_only=False)
                worksheet = workbook.active
                headers = [cell.value for cell in worksheet[1]]
                self.assertEqual(headers, ["SKU", "Название", "Бренд", "Продажи, шт", "Вес, кг", "Вес, кг (ед.)"])
                self.assertNotIn("__row_hash", headers)
                self.assertIsNone(worksheet.auto_filter.ref)
                self.assertIsInstance(worksheet["D2"].value, (int, float))
                self.assertEqual(worksheet["D2"].value, 5)
                self.assertIsInstance(worksheet["E2"].value, (int, float))
                self.assertIsInstance(worksheet["F2"].value, (int, float))
                self.assertEqual(worksheet["E2"].value, 1.5)
                self.assertEqual(worksheet["F2"].value, 0.5)

                downloaded = client.get("/api/exports/download-file", params={"path": str(xlsx_path)})
                self.assertEqual(downloaded.status_code, 200)

                with patch.object(
                    app.state.repository,
                    "fetch_export_products_dataframe",
                    side_effect=AssertionError("CSV export should use DuckDB COPY, not pandas chunks."),
                ):
                    csv_job = client.post(
                        "/api/exports/build-jobs",
                        json={
                            **preview_payload,
                            "excluded_row_hashes": [excluded_hash],
                            "output_dir": str(output_dir),
                            "confirm_large_export": False,
                            "export_format": "csv",
                        },
                    )
                    self.assertEqual(csv_job.status_code, 200)
                    csv_job_payload = csv_job.json()
                    self.assertIn(csv_job_payload["status"], {"queued", "running", "succeeded"})
                    csv_job_id = csv_job_payload["id"]
                    for _ in range(40):
                        csv_job_payload = client.get(f"/api/exports/build-jobs/{csv_job_id}").json()
                        if csv_job_payload["status"] == "succeeded":
                            break
                        time.sleep(0.05)
                self.assertEqual(csv_job_payload["status"], "succeeded")
                self.assertEqual(csv_job_payload["progress"], 100.0)
                csv_result = csv_job_payload["result"]
                self.assertEqual(csv_result["export_format"], "csv")
                self.assertEqual(csv_result["total"], 1)
                csv_path = Path(csv_result["artifacts"][0]["path"])
                self.assertEqual(csv_path.suffix, ".csv")
                self.assertEqual(csv_result["artifacts"][0]["format"], "csv")
                self.assertTrue(csv_path.exists())
                self.assertTrue(csv_path.read_bytes().startswith(b"\xef\xbb\xbf"))
                csv_text = csv_path.read_text(encoding="utf-8-sig")
                self.assertIn("SKU;Название;Бренд;Продажи, шт", csv_text.splitlines()[0])

                split = client.post(
                    "/api/exports/build",
                    json={
                        "project_name": "unit",
                        "category_keys": ["lemon-oz", "soap-wb"],
                        "period_from": "2025-01",
                        "period_to": "2025-02",
                        "selected_columns": ["SKU", "Название"],
                        "filters": [],
                        "excluded_row_hashes": [],
                        "sort_column": "SKU",
                        "sort_direction": "asc",
                        "split_by_category": True,
                        "output_dir": str(output_dir),
                        "confirm_large_export": False,
                    },
                )
                self.assertEqual(split.status_code, 200)
                split_payload = split.json()
                self.assertEqual(split_payload["total"], 3)
                self.assertEqual(len(split_payload["artifacts"]), 2)
                self.assertEqual(
                    sorted(item["category_key"] for item in split_payload["artifacts"]),
                    ["lemon-oz", "soap-wb"],
                )

    def test_large_category_reports_build_aggregated_excel(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)
            repository.ensure_ready()

            source_file = root / "lemon.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {
                            "Маркетплейс": "Ozon",
                            "Категория": "Лимонная кислота",
                            "SKU": "sku-1",
                            "Название": "лимон 1 кг",
                            "Бренд": "Brand A",
                            "Тип": "Кислота",
                            "Продажи, шт": 3,
                            "Выручка, руб": 30,
                            "Объем, кг": 1.5,
                        },
                        {
                            "Маркетплейс": "Ozon",
                            "Категория": "Лимонная кислота",
                            "SKU": "sku-2",
                            "Название": "лимон 2 кг",
                            "Бренд": "Brand A",
                            "Тип": "Кислота",
                            "Продажи, шт": 5,
                            "Выручка, руб": 50,
                            "Объем, кг": 2.5,
                        },
                    ]
                ),
                source_file,
            )
            repository.import_products_file_idempotent(
                run_id="run-report-1",
                csv_path=source_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="lemon-oz",
            )
            repository.upsert_cube_entry(
                {
                    "project_name": "unit",
                    "year": 2025,
                    "month": 1,
                    "marketplace": "Ozon",
                    "marketplace_code": "oz",
                    "category_key": "lemon-oz",
                    "category_name": "Лимонная кислота",
                    "rows_count": 1_200_000,
                    "days_loaded": 31,
                    "days_in_month": 31,
                    "data_actual_until": "2025-01-31",
                    "source_processed_file_path": str(source_file),
                    "file_hash": "hash",
                }
            )

            app = create_app(settings, start_workers=False)
            with TestClient(app) as client:
                options = client.get("/api/reports/options", params={"project_name": "unit"})
                self.assertEqual(options.status_code, 200)
                category = options.json()["categories"][0]
                self.assertTrue(category["is_heavy"])
                self.assertEqual(category["rows_count"], 1_200_000)
                self.assertIn("category_month", category["available_reports"])

                preview_payload = {
                    "project_name": "unit",
                    "report_type": "brand_month",
                    "category_keys": ["lemon-oz"],
                    "period_from": "2025-01",
                    "period_to": "2025-01",
                    "export_format": "xlsx",
                    "max_rows": 5000,
                    "limit": 100,
                    "offset": 0,
                }
                preview = client.post("/api/reports/preview", json=preview_payload)
                self.assertEqual(preview.status_code, 200)
                preview_payload_response = preview.json()
                self.assertEqual(preview_payload_response["total"], 1)
                self.assertEqual(preview_payload_response["rows"][0]["Бренд"], "Brand A")
                self.assertEqual(preview_payload_response["rows"][0]["Продажи, шт"], 8)
                self.assertEqual(preview_payload_response["rows"][0]["Выручка, руб"], 80)
                self.assertEqual(preview_payload_response["rows"][0]["Объем, кг"], 4)

                with patch.object(
                    app.state.repository,
                    "fetch_report_dataframe",
                    side_effect=AssertionError("Report XLSX export should use DuckDB COPY, not pandas/openpyxl."),
                ):
                    built = client.post(
                        "/api/reports/build",
                        json={**preview_payload, "output_dir": str(root / "reports")},
                    )
                self.assertEqual(built.status_code, 200)
                artifact = built.json()["artifacts"][0]
                xlsx_path = Path(artifact["path"])
                self.assertTrue(xlsx_path.exists())
                self.assertEqual(xlsx_path.parent, (root / "reports" / "unit").resolve())

                from openpyxl import load_workbook

                workbook = load_workbook(xlsx_path, read_only=False)
                worksheet = workbook.active
                headers = [cell.value for cell in worksheet[1]]
                self.assertIn("Бренд", headers)
                self.assertIn("Продажи, шт", headers)
                sales_column = headers.index("Продажи, шт") + 1
                self.assertIsInstance(worksheet.cell(row=2, column=sales_column).value, (int, float))
                downloaded = client.get("/api/reports/download-file", params={"path": str(xlsx_path)})
                self.assertEqual(downloaded.status_code, 200)

                cube = client.get("/api/workflow/pipeline/cube", params={"project_name": "unit"})
                self.assertEqual(cube.status_code, 200)
                cube_item = cube.json()["items"][0]
                self.assertTrue(cube_item["is_heavy"])
                self.assertEqual(cube_item["data_mode"], "heavy")
                self.assertIsNotNone(cube_item["reports_built_at"])

    def test_report_csv_build_uses_duckdb_copy_and_preserves_csv_contract(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)
            repository.ensure_ready()

            lemon_file = root / "lemon.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {
                            "Маркетплейс": "Ozon",
                            "Категория": "Лимонная кислота",
                            "SKU": "sku-1",
                            "Название": "лимон 1 кг",
                            "Бренд": "Brand A",
                            "Продажи, шт": 3,
                            "Выручка, руб": 30,
                            "Объем, кг": 1.5,
                        },
                        {
                            "Маркетплейс": "Ozon",
                            "Категория": "Лимонная кислота",
                            "SKU": "sku-2",
                            "Название": "лимон 2 кг",
                            "Бренд": "Brand A",
                            "Продажи, шт": 5,
                            "Выручка, руб": 50,
                            "Объем, кг": 2.5,
                        },
                    ]
                ),
                lemon_file,
            )
            repository.import_products_file_idempotent(
                run_id="run-report-csv-1",
                csv_path=lemon_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="lemon-oz",
            )
            soap_file = root / "soap.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {
                            "Маркетплейс": "Wildberries",
                            "Категория": "Мыло",
                            "SKU": "soap-1",
                            "Название": "мыло",
                            "Бренд": "Brand B",
                            "Продажи, шт": 9,
                            "Выручка, руб": 90,
                            "Объем, кг": 3,
                        }
                    ]
                ),
                soap_file,
            )
            repository.import_products_file_idempotent(
                run_id="run-report-csv-2",
                csv_path=soap_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=2,
                marketplace_code="wb",
                category_key="soap-wb",
            )

            app = create_app(settings, start_workers=False)
            with TestClient(app) as client, patch.object(
                app.state.repository,
                "fetch_report_dataframe",
                side_effect=AssertionError("Report CSV export should use DuckDB COPY, not pandas."),
            ):
                built = client.post(
                    "/api/reports/build",
                    json={
                        "project_name": "unit",
                        "report_type": "brand_month",
                        "category_keys": ["lemon-oz"],
                        "period_from": "2025-01",
                        "period_to": "2025-01",
                        "export_format": "csv",
                        "output_dir": str(root / "reports"),
                        "max_rows": 5000,
                    },
                )
                self.assertEqual(built.status_code, 200)

            payload = built.json()
            self.assertEqual(payload["total"], 1)
            self.assertEqual(payload["source_total"], 1)
            artifact = payload["artifacts"][0]
            csv_path = Path(artifact["path"])
            self.assertTrue(csv_path.exists())
            self.assertGreater(csv_path.stat().st_size, 0)
            self.assertEqual(csv_path.parent, (root / "reports" / "unit").resolve())
            self.assertTrue(csv_path.read_bytes().startswith(b"\xef\xbb\xbf"))
            csv_text = csv_path.read_text(encoding="utf-8-sig")
            self.assertIn("Период;Год;Месяц;Маркетплейс;Категория;Бренд", csv_text.splitlines()[0])
            rows = list(csv.DictReader(StringIO(csv_text), delimiter=";"))
            self.assertEqual(len(rows), 1)
            row = rows[0]
            self.assertEqual(row["Период"], "2025-01")
            self.assertEqual(row["Маркетплейс"], "Ozon")
            self.assertEqual(row["Категория"], "Лимонная кислота")
            self.assertEqual(row["Бренд"], "Brand A")
            self.assertEqual(float(row["Продажи, шт"].replace(",", ".")), 8.0)
            self.assertEqual(float(row["Выручка, руб"].replace(",", ".")), 80.0)
            self.assertNotIn("Мыло", csv_text)
            self.assertNotIn("2025-02", csv_text)

    def test_duckdb_csv_helper_writes_zero_rows_with_header_and_logs_sql_errors(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)
            repository.ensure_ready()

            target = root / "nested" / "zero.csv"
            result = repository.export_query_to_csv(
                "SELECT 1 AS id, 'лимон' AS name WHERE false",
                target,
                delimiter=";",
                header=True,
            )
            self.assertEqual(result.status, "success")
            self.assertEqual(result.row_count, 0)
            self.assertEqual(result.output_path, target)
            self.assertTrue(target.exists())
            self.assertTrue(target.read_bytes().startswith(b"\xef\xbb\xbf"))
            self.assertEqual(target.read_text(encoding="utf-8-sig").splitlines()[0], "id;name")

            named_target = root / "nested" / "named.csv"
            named_result = repository.export_query_to_csv(
                "SELECT $value AS id, $label AS name",
                named_target,
                params={"value": 7, "label": "семь"},
            )
            self.assertEqual(named_result.row_count, 1)
            named_rows = list(csv.DictReader(StringIO(named_target.read_text(encoding="utf-8-sig")), delimiter=";"))
            self.assertEqual(named_rows, [{"id": "7", "name": "семь"}])

            with self.assertLogs("mpstats_app.repositories.duckdb_repository", level="ERROR") as logs:
                with self.assertRaises(Exception):
                    repository.export_query_to_csv("SELECT missing_column FROM missing_table", root / "broken.csv")
            self.assertTrue(any("DuckDB COPY CSV export failed" in message for message in logs.output))

    def test_product_csv_export_preserves_date_column(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)

            source_file = root / "products.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {
                            "Дата": "01.01.2025",
                            "SKU": "sku-1",
                            "Продажи, шт": "5",
                            "Объем, кг": "1,5",
                        }
                    ]
                ),
                source_file,
            )
            repository.import_products_file_idempotent(
                run_id="run-1",
                csv_path=source_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="sugar",
            )

            target = root / "exports" / "products.csv"
            repository.export_products_to_csv(
                table_name=settings.products_table,
                target=target,
                project_name="unit",
                output_columns=["Дата", "SKU"],
                period_from_index=2025 * 12 + 1,
                period_to_index=2025 * 12 + 1,
            )

            rows = list(csv.DictReader(StringIO(target.read_text(encoding="utf-8-sig")), delimiter=";"))
            self.assertEqual(rows, [{"Дата": "01.01.2025", "SKU": "sku-1"}])

    def test_flat_csv_decimal_comma_preserves_protected_columns_and_xlsx_types(self) -> None:
        query = """
            SELECT
                '2025.01.02' AS " ДАТА ",
                'ABC.123' AS " Sku ",
                'seller.ru' AS "продавец",
                'cat.1' AS "категория",
                'Brand.1' AS "бренд",
                '2025.1' AS "год",
                '01.1' AS "месяц",
                'sub.1' AS "подкатегория",
                'type.1' AS "тип",
                sales AS "Продажи",
                price AS "Цена",
                revenue AS "Выручка",
                volume AS "Объем",
                new_metric AS "Новая метрика",
                site AS "Сайт",
                code AS "Артикул"
            FROM (
                VALUES
                    (1, true, 123.45, 0.5, CAST(1000.00 AS DECIMAL(10, 2)), CAST(NULL AS DOUBLE), 7.25, 'site.ru', 'A.B'),
                    (2, false, 999.9, 9.9, CAST(9.90 AS DECIMAL(10, 2)), 9.9, 9.9, 'skip.ru', 'Z.Z')
            ) AS rows(row_id, include_row, sales, price, revenue, volume, new_metric, site, code)
            WHERE include_row
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)
            repository.ensure_ready()

            csv_target = root / "flat" / "decimal-comma.csv"
            csv_result = repository.export_query_to_csv(query, csv_target)
            self.assertEqual(csv_result.status, "success")
            self.assertEqual(csv_result.row_count, 1)
            self.assertTrue(csv_target.read_bytes().startswith(b"\xef\xbb\xbf"))
            csv_text = csv_target.read_text(encoding="utf-8-sig")
            self.assertNotIn("skip.ru", csv_text)
            rows = list(csv.DictReader(StringIO(csv_text), delimiter=";"))
            self.assertEqual(len(rows), 1)
            row = rows[0]

            self.assertEqual(row[" ДАТА "], "2025.01.02")
            self.assertEqual(row[" Sku "], "ABC.123")
            self.assertEqual(row["продавец"], "seller.ru")
            self.assertEqual(row["категория"], "cat.1")
            self.assertEqual(row["бренд"], "Brand.1")
            self.assertEqual(row["год"], "2025.1")
            self.assertEqual(row["месяц"], "01.1")
            self.assertEqual(row["подкатегория"], "sub.1")
            self.assertEqual(row["тип"], "type.1")
            self.assertEqual(row["Продажи"], "123,45")
            self.assertEqual(row["Цена"], "0,5")
            self.assertNotIn(".", row["Выручка"])
            self.assertTrue(row["Выручка"].startswith("1000,"))
            self.assertEqual(row["Объем"], "")
            self.assertEqual(row["Новая метрика"], "7,25")
            self.assertEqual(row["Сайт"], "site.ru")
            self.assertEqual(row["Артикул"], "A.B")

            xlsx_target = root / "flat" / "decimal-comma.xlsx"
            xlsx_result = repository.export_query_to_xlsx(query, xlsx_target)
            self.assertEqual(xlsx_result.status, "success")
            self.assertEqual(xlsx_result.row_count, 1)

            from openpyxl import load_workbook

            workbook = load_workbook(xlsx_target, read_only=False)
            worksheet = workbook.active
            headers = [cell.value for cell in worksheet[1]]
            sales_column = headers.index("Продажи") + 1
            price_column = headers.index("Цена") + 1
            self.assertIsInstance(worksheet.cell(row=2, column=sales_column).value, (int, float))
            self.assertIsInstance(worksheet.cell(row=2, column=price_column).value, (int, float))
            self.assertEqual(worksheet.cell(row=2, column=sales_column).value, 123.45)

    def test_flat_xlsx_helper_uses_duckdb_copy_and_preserves_types(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)
            repository.ensure_ready()

            target = root / "flat" / "typed.xlsx"
            result = repository.export_query_to_xlsx(
                "SELECT 1 AS id, 2.5 AS amount, DATE '2025-01-02' AS sold_at",
                target,
                sheet_name="Data",
            )
            self.assertEqual(result.status, "success")
            self.assertEqual(result.format, "xlsx")
            self.assertEqual(result.row_count, 1)
            self.assertTrue(target.exists())

            from openpyxl import load_workbook

            workbook = load_workbook(target, read_only=False)
            worksheet = workbook.active
            self.assertEqual([cell.value for cell in worksheet[1]], ["id", "amount", "sold_at"])
            self.assertIsInstance(worksheet["A2"].value, int)
            self.assertIsInstance(worksheet["B2"].value, float)
            self.assertIsNotNone(worksheet["C2"].value)

    def test_flat_xlsx_helper_row_limit_and_openpyxl_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)
            repository.ensure_ready()

            too_large = root / "flat" / "too-large.xlsx"
            with self.assertRaisesRegex(ValueError, "XLSX row limit exceeded"):
                repository.export_query_to_xlsx("SELECT range AS id FROM range(1048577)", too_large)
            self.assertFalse(too_large.exists())

            fallback_target = root / "flat" / "fallback.xlsx"
            with patch.object(repository, "_load_excel_extension", side_effect=RuntimeError("extension unavailable")):
                result = repository.export_query_to_xlsx(
                    "SELECT 7 AS id, DATE '2025-01-03' AS sold_at",
                    fallback_target,
                )
            self.assertEqual(result.status, "fallback")
            self.assertIn("extension unavailable", result.error or "")
            self.assertEqual(result.row_count, 1)

            from openpyxl import load_workbook

            workbook = load_workbook(fallback_target, read_only=False)
            worksheet = workbook.active
            self.assertEqual([cell.value for cell in worksheet[1]], ["id", "sold_at"])
            self.assertEqual(worksheet["A2"].value, 7)

    def test_workflow_categories_classify_preview_and_save(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            app = create_app(settings, start_workers=False)

            input_file = root / "ready.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {
                            "Дата": "01.01.2025",
                            "Маркетплейс": "Ozon",
                            "Категория": "Лимонная кислота",
                            "SKU": "123",
                            "Бренд": "Brand",
                            "Название": "лимон тест",
                            "Продажи, шт": 5,
                        }
                    ]
                ),
                input_file,
            )

            with TestClient(app) as client:
                settings_response = client.put(
                    "/api/workflow/settings",
                    json={
                        "cookie": "session=test",
                        "api_token": "api-token-test",
                        "project_name": "unit",
                        "workflow_mode": "historical_backfill",
                        "start_year": 2024,
                        "start_month": 1,
                        "end_year": 2026,
                        "end_month": 5,
                    },
                )
                self.assertEqual(settings_response.status_code, 200)
                self.assertEqual(settings_response.json()["project_name"], "unit")
                self.assertEqual(settings_response.json()["api_token"], "api-token-test")
                self.assertEqual(settings_response.json()["start_year"], 2024)
                self.assertEqual(settings_response.json()["end_month"], 5)

                loaded_settings = client.get("/api/workflow/settings")
                self.assertEqual(loaded_settings.status_code, 200)
                self.assertEqual(loaded_settings.json()["workflow_mode"], "historical_backfill")
                self.assertEqual(loaded_settings.json()["api_token"], "api-token-test")
                self.assertEqual(loaded_settings.json()["start_month"], 1)

                categories = client.get("/api/workflow/categories")
                self.assertEqual(categories.status_code, 200)
                category_rows = categories.json()["categories"]
                self.assertGreaterEqual(len(category_rows), 3)
                self.assertFalse(any(row["category_name"] == "Пустой путь" for row in category_rows))
                lemon_oz = next(row for row in category_rows if row["category_name"] == "Лимонная кислота" and row["mp_code"] == "oz")
                lemon_wb = next(row for row in category_rows if row["category_name"] == "Лимонная кислота" and row["mp_code"] == "wb")
                self.assertTrue(lemon_oz["fbs"])
                self.assertEqual(lemon_oz["source_type"], "category")
                self.assertFalse(lemon_wb["fbs"])
                self.assertEqual(lemon_wb["source_type"], "category")
                yandex_category = next(row for row in category_rows if row["category_name"] == "Яндекс тест")
                self.assertFalse(yandex_category["fbs"])
                self.assertEqual(yandex_category["source_type"], "category")
                oil = next(row for row in category_rows if row["category_name"] == "Масло")
                self.assertIn("Подсолнечное", str(oil["filter_json"]))
                switched = [row for row in category_rows if row["category_name"] == "Смена пути"]
                self.assertEqual([row["period_from"] for row in switched], ["2025-01", "2025-02"])
                self.assertEqual([row["period_to"] for row in switched], ["2025-01", "2025-02"])

                source = client.get("/api/workflow/categories/source")
                self.assertEqual(source.status_code, 200)
                source_rows = source.json()["rows"]
                self.assertGreaterEqual(len(source_rows), 4)
                source_rows.append(
                    {
                        "active": True,
                        "category_name": "Новая категория",
                        "marketplace": "WB",
                        "fbs": True,
                        "source_type": "subject",
                        "period_from": "2025",
                        "period_to": "2025",
                        "comment": "unit",
                        "path": "Продукты/Новая",
                        "filter_text": "\"мыло\"&NOT\"хозяйственное\"",
                        "path2": "",
                        "filter2_text": "",
                        "actualization": "",
                    }
                )
                source_rows.append(
                    {
                        "active": True,
                        "category_name": "WB по умолчанию",
                        "marketplace": "WB",
                        "source_type": "category",
                        "period_from": "2025",
                        "period_to": "2025",
                        "comment": "unit",
                        "path": "Продукты/WB default",
                        "filter_text": "",
                        "path2": "",
                        "filter2_text": "",
                        "actualization": "",
                    }
                )
                source_rows.append(
                    {
                        "active": True,
                        "category_name": "Яндекс из UI",
                        "marketplace": "ЯМ",
                        "fbs": True,
                        "source_type": "subject",
                        "period_from": "2025",
                        "period_to": "2025",
                        "comment": "unit",
                        "path": "Продукты/YM",
                        "filter_text": "",
                        "path2": "",
                        "filter2_text": "",
                        "actualization": "",
                    }
                )
                saved_source = client.put("/api/workflow/categories/source", json={"rows": source_rows})
                self.assertEqual(saved_source.status_code, 200)
                refreshed_categories = client.get("/api/workflow/categories").json()["categories"]
                self.assertTrue(any(row["category_name"] == "Новая категория" for row in refreshed_categories))
                new_category = next(row for row in refreshed_categories if row["category_name"] == "Новая категория")
                self.assertTrue(new_category["fbs"])
                self.assertEqual(new_category["source_type"], "subject")
                wb_default = next(row for row in refreshed_categories if row["category_name"] == "WB по умолчанию")
                self.assertTrue(wb_default["fbs"])
                self.assertEqual(wb_default["source_type"], "category")
                ym_from_ui = next(row for row in refreshed_categories if row["category_name"] == "Яндекс из UI")
                self.assertFalse(ym_from_ui["fbs"])
                self.assertEqual(ym_from_ui["source_type"], "category")
                saved_header = (root / "Справочник категорий MP STATS.csv").read_text(encoding="utf-8-sig").splitlines()[0]
                self.assertIn("Тип выгрузки", saved_header)
                saved_rows = saved_source.json()["rows"]
                saved_new_row = next(row for row in saved_rows if row["category_name"] == "Новая категория")
                self.assertEqual(saved_new_row["source_type"], "subject")
                new_filter = json.loads(new_category["filter_json"])
                self.assertEqual(new_filter["name"]["operator"], "AND")
                self.assertEqual(new_filter["name"]["condition1"]["type"], "contains")
                self.assertEqual(new_filter["name"]["condition2"]["type"], "notContains")

                classified = client.post(
                    "/api/workflow/classify",
                    json={"project_name": "unit", "input_file": str(input_file), "overwrite_input": False},
                )
                self.assertEqual(classified.status_code, 200)
                preview = classified.json()["preview"]
                self.assertEqual(preview["total"], 1)
                self.assertEqual(preview["rows"][0]["Тип"], "Кислота")

                uploaded = client.post(
                    "/api/workflow/classify-upload",
                    params={"project_name": "unit", "filename": "external.csv", "write_xlsx": False},
                    content="Название,SKU\nлимон внешний,456\n".encode("utf-8"),
                    headers={"content-type": "text/csv"},
                )
                self.assertEqual(uploaded.status_code, 200)
                uploaded_payload = uploaded.json()
                self.assertNotEqual(uploaded_payload["input_file"], uploaded_payload["output_file"])
                self.assertTrue(Path(uploaded_payload["output_file"]).exists())
                self.assertIn("external_classification", uploaded_payload["output_file"])
                self.assertEqual(uploaded_payload["preview"]["total"], 1)
                self.assertEqual(uploaded_payload["preview"]["rows"][0]["Тип"], "Кислота")
                downloaded = client.get("/api/workflow/download-file", params={"path": uploaded_payload["output_file"]})
                self.assertEqual(downloaded.status_code, 200)
                self.assertIn("Кислота", downloaded.text)

                if _has_openpyxl():
                    xlsx_buffer = BytesIO()
                    pd.DataFrame([{"Название": "лимон excel", "SKU": "789"}]).to_excel(xlsx_buffer, index=False)
                    uploaded_xlsx = client.post(
                        "/api/workflow/classify-upload",
                        params={"project_name": "unit", "filename": "external.xlsx", "write_xlsx": False},
                        content=xlsx_buffer.getvalue(),
                        headers={"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
                    )
                    self.assertEqual(uploaded_xlsx.status_code, 200)
                    self.assertEqual(uploaded_xlsx.json()["preview"]["rows"][0]["Тип"], "Кислота")

                saved = client.post(
                    "/api/workflow/save-to-db",
                    json={"project_name": "unit", "file_path": classified.json()["output_file"]},
                )
                self.assertEqual(saved.status_code, 200)
                self.assertEqual(saved.json()["rows"], 1)

    def test_workflow_categories_follow_global_csv_after_project_switch(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            app = create_app(settings, start_workers=False)
            source = root / "Справочник категорий MP STATS.csv"

            with TestClient(app) as client:
                initial = client.get("/api/workflow/categories")
                self.assertEqual(initial.status_code, 200)
                self.assertTrue(
                    any(row["category_name"] == "Лимонная кислота" for row in initial.json()["categories"])
                )

                source.write_text(
                    "\n".join(
                        [
                            "Чек;Категория;МП;FBS;От;До;Комментарий;Путь;Фильтр;Путь2;Фильтр2;Актуализация",
                            ";Локальный второй комп;WB;1;2025;2025;;Продукты/Локальный;;;;",
                        ]
                    )
                    + "\n",
                    encoding="utf-8-sig",
                )
                settings_response = client.put(
                    "/api/workflow/settings",
                    json={
                        "cookie": "",
                        "api_token": "",
                        "project_name": "other-project",
                        "workflow_mode": "historical_backfill",
                        "start_year": 2025,
                        "start_month": 1,
                        "end_year": 2025,
                        "end_month": 1,
                    },
                )
                self.assertEqual(settings_response.status_code, 200)

                refreshed = client.get("/api/workflow/categories")
                self.assertEqual(refreshed.status_code, 200)
                refreshed_names = {row["category_name"] for row in refreshed.json()["categories"]}
                self.assertIn("Локальный второй комп", refreshed_names)
                self.assertNotIn("Лимонная кислота", refreshed_names)

    def test_smart_plan_compares_expected_tasks_with_local_files(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            app = create_app(settings, start_workers=False)

            with TestClient(app) as client:
                categories = client.get("/api/workflow/categories").json()["categories"]
                category_ids = [
                    row["category_id"]
                    for row in categories
                    if row["category_name"] in {"Лимонная кислота", "Масло"}
                ]
                plan_response = client.post(
                    "/api/workflow/pipeline/plans",
                    json={
                        "project_name": "smart-plan-unit",
                        "run_type": "historical_backfill",
                        "category_ids": category_ids,
                        "start_year": 2025,
                        "start_month": 1,
                        "end_year": 2025,
                        "end_month": 2,
                        "settings": {
                            "overwrite_raw": False,
                            "overwrite_processed": False,
                            "overwrite_db": False,
                            "max_parallel_downloads": 1,
                            "retry_count": 0,
                            "timeout_seconds": 300,
                            "pause_between_requests": 0,
                            "max_weight_kg": 40,
                        },
                    },
                )
                self.assertEqual(plan_response.status_code, 200)
                run_id = plan_response.json()["id"]
                tasks = client.get(f"/api/workflow/pipeline/runs/{run_id}/tasks").json()["tasks"]
                self.assertEqual(len(tasks), 6)

                write_semicolon_csv(pd.DataFrame([{"SKU": "raw-only"}]), Path(tasks[1]["raw_file_path"]))
                write_semicolon_csv(pd.DataFrame([{"SKU": "ready"}]), Path(tasks[2]["processed_file_path"]))
                write_semicolon_csv(pd.DataFrame([{"SKU": "ready"}]), Path(tasks[2]["classified_file_path"]))
                write_semicolon_csv(pd.DataFrame([{"SKU": "old-processed"}]), Path(tasks[3]["processed_file_path"]))
                write_semicolon_csv(pd.DataFrame([{"SKU": "new-raw"}]), Path(tasks[3]["raw_file_path"]))
                old_time = 1_700_000_000
                new_time = old_time + 100
                os.utime(tasks[3]["processed_file_path"], (old_time, old_time))
                os.utime(tasks[3]["raw_file_path"], (new_time, new_time))
                app.state.repository.update_download_task(tasks[4]["id"], {"status": "failed", "error_message": "unit failure"})

                smart_plan = client.get(f"/api/workflow/pipeline/runs/{run_id}/smart-plan")
                self.assertEqual(smart_plan.status_code, 200)
                summary = smart_plan.json()["summary"]
                self.assertEqual(summary["total"], 6)
                self.assertEqual(summary["missing"], 2)
                self.assertEqual(summary["incomplete"], 1)
                self.assertEqual(summary["ready"], 1)
                self.assertEqual(summary["stale"], 1)
                self.assertEqual(summary["failed"], 1)
                self.assertEqual(smart_plan.json()["recommended_action"]["key"], "retry_failed")

                ready_only = client.get(f"/api/workflow/pipeline/runs/{run_id}/smart-plan", params={"status": "ready"})
                self.assertEqual(ready_only.status_code, 200)
                self.assertEqual(len(ready_only.json()["tasks"]), 1)
                self.assertEqual(ready_only.json()["tasks"][0]["recommended_action"], "Собрать куб из готовых файлов")

    def test_smart_plan_batches_cube_lookup_and_skips_saved_file_hash(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            app = create_app(settings, start_workers=False)

            with TestClient(app) as client:
                category_id = client.get("/api/workflow/categories").json()["categories"][0]["category_id"]
                plan_response = client.post(
                    "/api/workflow/pipeline/plans",
                    json={
                        "project_name": "smart-plan-unit",
                        "run_type": "historical_backfill",
                        "category_ids": [category_id],
                        "start_year": 2025,
                        "start_month": 1,
                        "end_year": 2025,
                        "end_month": 1,
                        "settings": {
                            "overwrite_raw": False,
                            "overwrite_processed": False,
                            "overwrite_db": False,
                            "max_parallel_downloads": 1,
                            "retry_count": 0,
                            "timeout_seconds": 300,
                            "pause_between_requests": 0,
                            "max_weight_kg": 40,
                        },
                    },
                )
                self.assertEqual(plan_response.status_code, 200)
                run_id = plan_response.json()["id"]
                task = client.get(f"/api/workflow/pipeline/runs/{run_id}/tasks").json()["tasks"][0]
                classified_path = write_semicolon_csv(pd.DataFrame([{"SKU": "ready"}]), Path(task["classified_file_path"]))
                old_time = 1_700_000_000
                os.utime(classified_path, (old_time, old_time))
                app.state.repository.upsert_cube_entry(
                    {
                        "project_name": task["project_name"],
                        "year": task["year"],
                        "month": task["month"],
                        "marketplace": task["marketplace"],
                        "marketplace_code": task["marketplace_code"],
                        "category_key": task["category_key"],
                        "category_name": task["category_name"],
                        "rows_count": 1,
                        "source_processed_file_path": str(classified_path),
                        "file_hash": "not-the-current-file-hash",
                    }
                )

                with (
                    patch.object(app.state.repository, "get_cube_entry", side_effect=AssertionError("smart-plan must batch cube lookups")),
                    patch.object(app.state.smart_plan_service, "_file_sha1", side_effect=AssertionError("saved file hash should not be recalculated")),
                ):
                    smart_plan = client.get(f"/api/workflow/pipeline/runs/{run_id}/smart-plan")

                self.assertEqual(smart_plan.status_code, 200)
                payload = smart_plan.json()
                self.assertEqual(payload["summary"]["ready"], 1)
                self.assertEqual(payload["summary"]["saved_to_db"], 1)
                self.assertTrue(payload["tasks"][0]["has_cube"])
                self.assertEqual(payload["tasks"][0]["smart_status"], "ready")

    def test_smart_pipeline_plan_rebuild_dedup_retry_and_monthly_sync(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            app = create_app(settings, start_workers=False)

            with TestClient(app) as client:
                categories = client.get("/api/workflow/categories").json()["categories"]
                category_ids = [row["category_id"] for row in categories if row["category_name"] == "Лимонная кислота"][:2]

                plan_response = client.post(
                    "/api/workflow/pipeline/plans",
                    json={
                        "project_name": "unit",
                        "run_type": "historical_backfill",
                        "category_ids": category_ids,
                        "start_year": 2025,
                        "start_month": 1,
                        "end_year": 2025,
                        "end_month": 2,
                        "settings": {
                            "overwrite_raw": False,
                            "overwrite_processed": False,
                            "overwrite_db": False,
                            "max_parallel_downloads": 1,
                            "retry_count": 0,
                            "timeout_seconds": 300,
                            "pause_between_requests": 0,
                            "max_weight_kg": 40,
                        },
                    },
                )
                self.assertEqual(plan_response.status_code, 200)
                run_id = plan_response.json()["id"]
                tasks_response = client.get(f"/api/workflow/pipeline/runs/{run_id}/tasks")
                self.assertEqual(tasks_response.status_code, 200)
                tasks = tasks_response.json()["tasks"]
                self.assertEqual(len(tasks), 4)
                self.assertEqual({task["source_type"] for task in tasks}, {"category"})
                smart_pipeline = app.state.smart_pipeline_service
                oz_task = next(task for task in tasks if task["marketplace_code"] == "oz")
                wb_task = next(task for task in tasks if task["marketplace_code"] == "wb")
                self.assertEqual(smart_pipeline._task_to_export_task(oz_task).get("fbs"), 1)
                self.assertNotIn("fbs", smart_pipeline._task_to_export_task(wb_task))

                switched_ids = [row["category_id"] for row in categories if row["category_name"] == "Смена пути"]
                switched_plan = client.post(
                    "/api/workflow/pipeline/plans",
                    json={
                        "project_name": "period-unit",
                        "run_type": "historical_backfill",
                        "category_ids": switched_ids,
                        "start_year": 2025,
                        "start_month": 1,
                        "end_year": 2025,
                        "end_month": 2,
                        "settings": {
                            "overwrite_raw": False,
                            "overwrite_processed": False,
                            "overwrite_db": False,
                            "max_parallel_downloads": 1,
                            "retry_count": 0,
                            "timeout_seconds": 300,
                            "pause_between_requests": 0,
                            "max_weight_kg": 40,
                        },
                    },
                )
                self.assertEqual(switched_plan.status_code, 200)
                switched_tasks = client.get(f"/api/workflow/pipeline/runs/{switched_plan.json()['id']}/tasks").json()["tasks"]
                self.assertEqual([(task["year"], task["month"], task["category_path"]) for task in switched_tasks], [
                    (2025, 1, "Продукты/Старый путь"),
                    (2025, 2, "Продукты/Новый путь"),
                ])

                for index, task in enumerate(tasks):
                    write_semicolon_csv(
                        pd.DataFrame(
                            [
                                {
                                    "Маркетплейс": task["marketplace"],
                                    "Категория": task["category_name"],
                                    "Артикул": f"sku-{index}",
                                    "SKU": f"лимон исправленный {index}",
                                    "Бренд": "brand",
                                    "Продажи, шт": 5,
                                    "Средняя цена, руб": 10,
                                    "Выручка, руб": 50,
                                }
                            ]
                        ),
                        Path(task["processed_file_path"]),
                    )
                    write_semicolon_csv(
                        pd.DataFrame(
                            [
                                {
                                    "Маркетплейс": task["marketplace"],
                                    "Категория": task["category_name"],
                                    "Артикул": f"sku-{index}",
                                    "SKU": f"лимон тест {index}",
                                    "Бренд": "brand",
                                    "Тип": "Старая классификация",
                                    "Продажи, шт": 5,
                                    "Средняя цена, руб": 10,
                                    "Выручка, руб": 50,
                                }
                            ]
                        ),
                        Path(task["classified_file_path"]),
                    )

                rebuilt = client.post(f"/api/workflow/pipeline/runs/{run_id}/rebuild-cube", json={"wait": True})
                self.assertEqual(rebuilt.status_code, 200)
                self.assertEqual(rebuilt.json()["completed_tasks"], 4)
                self.assertEqual(rebuilt.json()["failed_tasks"], 0)

                search = client.get("/api/products", params={"query": "лимон", "limit": 100})
                self.assertEqual(search.status_code, 200)
                self.assertEqual(search.json()["total"], 4)

                rebuilt_again = client.post(f"/api/workflow/pipeline/runs/{run_id}/rebuild-cube", json={"wait": True})
                self.assertEqual(rebuilt_again.status_code, 200)
                search_again = client.get("/api/products", params={"query": "лимон", "limit": 100})
                self.assertEqual(search_again.json()["total"], 4)

                reclassified = client.post(f"/api/workflow/pipeline/runs/{run_id}/reclassify-cube", json={"wait": True})
                self.assertEqual(reclassified.status_code, 200)
                reclassified_payload = reclassified.json()
                self.assertEqual(reclassified_payload["completed_tasks"], 4)
                self.assertEqual(reclassified_payload["failed_tasks"], 0)
                self.assertEqual(
                    reclassified_payload["operation_progress"],
                    {
                        "kind": "reclassify",
                        "total_files": 4,
                        "completed_files": 4,
                        "failed_files": 0,
                        "remaining_files": 0,
                        "progress": 100.0,
                        "status": "succeeded",
                    },
                )
                corrected_search = client.get("/api/products", params={"query": "исправленный", "limit": 100})
                self.assertEqual(corrected_search.status_code, 200)
                corrected_payload = corrected_search.json()
                self.assertEqual(corrected_payload["total"], 4)
                self.assertEqual({row["Тип"] for row in corrected_payload["rows"]}, {"Кислота"})
                old_search = client.get("/api/products", params={"query": "тест", "limit": 100})
                self.assertEqual(old_search.json()["total"], 0)

                for index, task in enumerate(tasks):
                    write_semicolon_csv(
                        pd.DataFrame(
                            [
                                {
                                    "SKU": f"raw-{index}",
                                    "Brand": "brand",
                                    "Name": f"лимон сырой {index} 3 x 175 г",
                                    "Sales": 5,
                                    "Average price": 100,
                                    "Revenue": 500,
                                    "Seller": "seller",
                                }
                            ]
                        ),
                        Path(task["raw_file_path"]),
                    )
                    write_semicolon_csv(pd.DataFrame([{"SKU": "старый processed"}]), Path(task["processed_file_path"]))
                    write_semicolon_csv(pd.DataFrame([{"SKU": "старый classified"}]), Path(task["classified_file_path"]))

                reprocessed = client.post(f"/api/workflow/pipeline/runs/{run_id}/reprocess-sources", json={"wait": True})
                self.assertEqual(reprocessed.status_code, 200)
                reprocessed_payload = reprocessed.json()
                self.assertEqual(reprocessed_payload["completed_tasks"], 4)
                self.assertEqual(reprocessed_payload["failed_tasks"], 0)
                self.assertEqual(
                    reprocessed_payload["operation_progress"],
                    {
                        "kind": "reprocess",
                        "total_files": 4,
                        "completed_files": 4,
                        "failed_files": 0,
                        "remaining_files": 0,
                        "progress": 100.0,
                        "status": "succeeded",
                    },
                )
                reprocessed_task = client.get(f"/api/workflow/pipeline/runs/{run_id}/tasks").json()["tasks"][0]
                processed_text = Path(reprocessed_task["processed_file_path"]).read_text(encoding="utf-8-sig")
                classified_text = Path(reprocessed_task["classified_file_path"]).read_text(encoding="utf-8-sig")
                self.assertIn("лимон сырой", processed_text)
                self.assertIn("Вес, кг", processed_text)
                self.assertIn("Вес, кг (ед.)", processed_text)
                self.assertNotIn("старый processed", processed_text)
                self.assertIn("лимон сырой", classified_text)
                self.assertNotIn("старый classified", classified_text)
                raw_search = client.get("/api/products", params={"query": "сырой", "limit": 100})
                self.assertEqual(raw_search.status_code, 200)
                raw_payload = raw_search.json()
                self.assertEqual(raw_payload["total"], 4)
                self.assertEqual({row["Тип"] for row in raw_payload["rows"]}, {"Кислота"})
                self.assertEqual(client.get("/api/products", params={"query": "исправленный", "limit": 100}).json()["total"], 0)

                repository: DuckDbAppRepository = app.state.repository
                repository.update_download_task(tasks[0]["id"], {"status": "failed", "error_message": "unit failure"})
                retry_errors = client.post(f"/api/workflow/pipeline/runs/{run_id}/retry-errors", json={"wait": True})
                self.assertEqual(retry_errors.status_code, 200)
                retried_error_task = client.get(f"/api/workflow/pipeline/runs/{run_id}/tasks").json()["tasks"][0]
                self.assertEqual(retried_error_task["status"], "saved_to_db")

                repository.update_download_task(tasks[1]["id"], {"status": "failed", "error_message": "unit failure"})
                retry = client.post(f"/api/workflow/pipeline/tasks/{tasks[1]['id']}/retry", json={"wait": True})
                self.assertEqual(retry.status_code, 200)
                retried_task = next(
                    task for task in client.get(f"/api/workflow/pipeline/runs/{run_id}/tasks").json()["tasks"]
                    if task["id"] == tasks[1]["id"]
                )
                self.assertEqual(retried_task["status"], "saved_to_db")

                monthly = client.post(
                    "/api/workflow/pipeline/monthly-sync",
                    json={
                        "project_name": "unit",
                        "start_immediately": False,
                        "wait": False,
                        "settings": {
                            "overwrite_raw": False,
                            "overwrite_processed": False,
                            "overwrite_db": False,
                            "max_parallel_downloads": 1,
                            "retry_count": 0,
                            "timeout_seconds": 300,
                            "pause_between_requests": 0,
                            "max_weight_kg": 40,
                        },
                    },
                )
                self.assertEqual(monthly.status_code, 200)
                self.assertEqual(monthly.json()["period_from"], "2025-03")
                self.assertEqual(monthly.json()["period_to"], "2025-03")

    def test_smart_pipeline_pause_and_stop_interrupt_background_run(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            app = create_app(settings, start_workers=False)

            with TestClient(app) as client:
                categories = client.get("/api/workflow/categories").json()["categories"]
                category_ids = [row["category_id"] for row in categories if row["category_name"] == "Лимонная кислота"][:2]

                def create_ready_plan(project_name: str) -> tuple[str, list[dict[str, object]]]:
                    plan_response = client.post(
                        "/api/workflow/pipeline/plans",
                        json={
                            "project_name": project_name,
                            "run_type": "historical_backfill",
                            "category_ids": category_ids,
                            "start_year": 2025,
                            "start_month": 1,
                            "end_year": 2025,
                            "end_month": 1,
                            "settings": {
                                "overwrite_raw": False,
                                "overwrite_processed": False,
                                "overwrite_db": False,
                                "max_parallel_downloads": 1,
                                "retry_count": 0,
                                "timeout_seconds": 300,
                                "pause_between_requests": 0,
                                "max_weight_kg": 40,
                            },
                        },
                    )
                    self.assertEqual(plan_response.status_code, 200)
                    run_id = plan_response.json()["id"]
                    tasks = client.get(f"/api/workflow/pipeline/runs/{run_id}/tasks").json()["tasks"]
                    for index, task in enumerate(tasks):
                        write_semicolon_csv(pd.DataFrame([{"SKU": f"готовый {project_name} {index}"}]), Path(str(task["classified_file_path"])))
                    return run_id, tasks

                pause_run_id, pause_tasks = create_ready_plan("pause-unit")
                service = app.state.smart_pipeline_service
                original_classify = service._classify_task
                pause_started = Event()
                pause_release = Event()

                def classify_and_wait_for_pause(task_id: str, *, settings: dict[str, object], force_reclassify: bool = False) -> None:
                    original_classify(task_id, settings=settings, force_reclassify=force_reclassify)
                    if task_id == pause_tasks[0]["id"]:
                        pause_started.set()
                        self.assertTrue(pause_release.wait(2.0))

                with patch.object(service, "_classify_task", side_effect=classify_and_wait_for_pause):
                    started = client.post(f"/api/workflow/pipeline/runs/{pause_run_id}/rebuild-cube", json={"wait": False})
                    self.assertEqual(started.status_code, 200)
                    self.assertTrue(pause_started.wait(2.0))
                    paused_request = client.post(f"/api/workflow/pipeline/runs/{pause_run_id}/pause")
                    self.assertEqual(paused_request.status_code, 200)
                    self.assertEqual(paused_request.json()["status"], "pausing")
                    pause_release.set()

                    paused_payload = {}
                    for _ in range(40):
                        paused_payload = client.get(f"/api/workflow/pipeline/runs/{pause_run_id}").json()
                        if paused_payload["status"] == "paused":
                            break
                        time.sleep(0.05)
                    self.assertEqual(paused_payload["status"], "paused")

                paused_tasks = client.get(f"/api/workflow/pipeline/runs/{pause_run_id}/tasks").json()["tasks"]
                self.assertEqual(paused_tasks[0]["status"], "classified")
                self.assertEqual(paused_tasks[0]["save_status"], "pending")
                self.assertEqual(paused_tasks[1]["status"], "pending")

                stop_run_id, stop_tasks = create_ready_plan("stop-unit")
                stop_started = Event()
                stop_release = Event()

                def classify_and_wait_for_stop(task_id: str, *, settings: dict[str, object], force_reclassify: bool = False) -> None:
                    original_classify(task_id, settings=settings, force_reclassify=force_reclassify)
                    if task_id == stop_tasks[0]["id"]:
                        stop_started.set()
                        self.assertTrue(stop_release.wait(2.0))

                with patch.object(service, "_classify_task", side_effect=classify_and_wait_for_stop):
                    started = client.post(f"/api/workflow/pipeline/runs/{stop_run_id}/rebuild-cube", json={"wait": False})
                    self.assertEqual(started.status_code, 200)
                    self.assertTrue(stop_started.wait(2.0))
                    stopped_request = client.post(f"/api/workflow/pipeline/runs/{stop_run_id}/stop")
                    self.assertEqual(stopped_request.status_code, 200)
                    self.assertEqual(stopped_request.json()["status"], "stopping")
                    self.assertTrue(app.state.repository.is_pipeline_stop_requested(stop_run_id))
                    stop_release.set()

                    stopped_payload = {}
                    for _ in range(40):
                        stopped_payload = client.get(f"/api/workflow/pipeline/runs/{stop_run_id}").json()
                        if stopped_payload["status"] == "stopped":
                            break
                        time.sleep(0.05)
                    self.assertEqual(stopped_payload["status"], "stopped")
                    self.assertEqual(stopped_payload["current_step"], "Остановлено")

                stopped_tasks = client.get(f"/api/workflow/pipeline/runs/{stop_run_id}/tasks").json()["tasks"]
                self.assertEqual(stopped_tasks[0]["status"], "classified")
                self.assertEqual(stopped_tasks[0]["save_status"], "pending")
                self.assertEqual(stopped_tasks[1]["status"], "pending")

    def test_smart_pipeline_deletes_cube_file_and_plan(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            app = create_app(settings, start_workers=False)

            with TestClient(app) as client:
                categories = client.get("/api/workflow/categories").json()["categories"]
                category_id = next(row["category_id"] for row in categories if row["category_name"] == "Лимонная кислота")
                plan_response = client.post(
                    "/api/workflow/pipeline/plans",
                    json={
                        "project_name": "unit",
                        "run_type": "historical_backfill",
                        "category_ids": [category_id],
                        "start_year": 2025,
                        "start_month": 1,
                        "end_year": 2025,
                        "end_month": 1,
                        "settings": {
                            "overwrite_raw": False,
                            "overwrite_processed": False,
                            "overwrite_db": False,
                            "max_parallel_downloads": 1,
                            "retry_count": 0,
                            "timeout_seconds": 300,
                            "pause_between_requests": 0,
                            "max_weight_kg": 40,
                        },
                    },
                )
                self.assertEqual(plan_response.status_code, 200)
                run_id = plan_response.json()["id"]
                task = client.get(f"/api/workflow/pipeline/runs/{run_id}/tasks").json()["tasks"][0]
                processed_path = Path(task["processed_file_path"])
                classified_path = Path(task["classified_file_path"])
                processed_path.parent.mkdir(parents=True, exist_ok=True)
                source_frame = pd.DataFrame(
                    [
                        {
                            "Маркетплейс": task["marketplace"],
                            "Категория": task["category_name"],
                            "Артикул": "sku-1",
                            "SKU": "лимон удаление",
                            "Бренд": "brand",
                            "Тип": "Кислота",
                            "Продажи, шт": 5,
                        }
                    ]
                )
                write_semicolon_csv(source_frame, processed_path)
                write_semicolon_csv(source_frame, classified_path)

                rebuilt = client.post(f"/api/workflow/pipeline/runs/{run_id}/rebuild-cube", json={"wait": True})
                self.assertEqual(rebuilt.status_code, 200)
                self.assertEqual(rebuilt.json()["completed_tasks"], 1)
                cube_entry = client.get("/api/workflow/pipeline/cube", params={"project_name": "unit"}).json()["items"][0]
                search = client.get("/api/products", params={"project_name": "unit", "query": "удаление", "limit": 100})
                self.assertEqual(search.json()["total"], 1)

                deleted_cube = client.delete(f"/api/workflow/pipeline/cube/{cube_entry['id']}")
                self.assertEqual(deleted_cube.status_code, 200)
                self.assertEqual(deleted_cube.json()["deleted"]["cube_registry"], 1)
                self.assertEqual(deleted_cube.json()["deleted"]["product_rows"], 1)
                self.assertEqual(client.get("/api/workflow/pipeline/cube", params={"project_name": "unit"}).json()["items"], [])
                task_after_cube_delete = client.get(f"/api/workflow/pipeline/runs/{run_id}/tasks").json()["tasks"][0]
                self.assertEqual(task_after_cube_delete["status"], "classified")
                self.assertEqual(task_after_cube_delete["save_status"], "pending")
                search_after_cube_delete = client.get("/api/products", params={"project_name": "unit", "query": "удаление", "limit": 100})
                self.assertEqual(search_after_cube_delete.json()["total"], 0)

                rebuilt_again = client.post(f"/api/workflow/pipeline/runs/{run_id}/rebuild-cube", json={"wait": True})
                self.assertEqual(rebuilt_again.status_code, 200)
                self.assertEqual(rebuilt_again.json()["completed_tasks"], 1)
                deleted_file = client.delete(
                    "/api/workflow/pipeline/files",
                    params={"project_name": "unit", "path": str(classified_path), "delete_cube": True},
                )
                self.assertEqual(deleted_file.status_code, 200)
                self.assertFalse(classified_path.exists())
                self.assertEqual(deleted_file.json()["deleted"]["files"], 1)
                self.assertEqual(deleted_file.json()["deleted"]["download_tasks"], 1)
                self.assertEqual(deleted_file.json()["cube_deletions"][0]["deleted"]["cube_registry"], 1)
                self.assertEqual(client.get("/api/workflow/pipeline/cube", params={"project_name": "unit"}).json()["items"], [])
                task_after_file_delete = client.get(f"/api/workflow/pipeline/runs/{run_id}/tasks").json()["tasks"][0]
                self.assertEqual(task_after_file_delete["classify_status"], "pending")
                self.assertEqual(task_after_file_delete["save_status"], "pending")

                deleted_plan = client.delete(f"/api/workflow/pipeline/runs/{run_id}")
                self.assertEqual(deleted_plan.status_code, 200)
                self.assertEqual(deleted_plan.json()["deleted"]["pipeline_runs"], 1)
                self.assertEqual(deleted_plan.json()["deleted"]["download_tasks"], 1)
                runs = client.get("/api/workflow/pipeline/runs", params={"project_name": "unit"}).json()["runs"]
                self.assertNotIn(run_id, {row["id"] for row in runs})

    def test_smart_pipeline_files_classify_kinds_and_hide_merged(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            project_root = root / "data" / "projects" / "unit"
            files = {
                "raw/2025-01/oz/sugar.csv": "raw",
                "processed/2025-01/oz/sugar.csv": "processed",
                "processed/2025-01/oz/sugar_classified.csv": "classified",
                "exports/MPStats_unit.xlsx": "export",
                "merged/cube_unit_2025_01.csv": "merged",
            }
            for relative_path in files:
                path = project_root / relative_path
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_text("sku\n1\n", encoding="utf-8")

            settings = make_settings(root)
            app = create_app(settings, start_workers=False)

            with TestClient(app) as client:
                response = client.get("/api/workflow/pipeline/files", params={"project_name": "unit"})

            self.assertEqual(response.status_code, 200)
            rows = response.json()["files"]
            by_relative_path = {row["relative_path"]: row for row in rows}
            self.assertEqual(by_relative_path["raw/2025-01/oz/sugar.csv"]["kind"], "raw")
            self.assertEqual(by_relative_path["processed/2025-01/oz/sugar.csv"]["kind"], "processed")
            self.assertEqual(by_relative_path["processed/2025-01/oz/sugar_classified.csv"]["kind"], "classified")
            self.assertEqual(by_relative_path["exports/MPStats_unit.xlsx"]["kind"], "export")
            self.assertNotIn("merged/cube_unit_2025_01.csv", by_relative_path)

    def test_quality_projects_report_and_missing_project(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            seed_project(root)
            settings = make_settings(root)
            repository = DuckDbAppRepository(settings)
            cube_file = root / "quality_cube.csv"
            write_semicolon_csv(
                pd.DataFrame(
                    [
                        {"Дата": "01.01.2025", "Название": "Лимон 1 кг", "Маркетплейс": "Ozon", "Категория": "Кислота", "SKU": "sku-1", "Продажи, шт": 10, "Средняя цена, руб": 10, "Выручка, руб": 100, "Объем, кг": 1.0},
                        {"Дата": "01.01.2025", "Название": "Лимон 2 кг", "Маркетплейс": "Ozon", "Категория": "Кислота", "SKU": "sku-2", "Продажи, шт": 12, "Средняя цена, руб": 10, "Выручка, руб": 120, "Объем, кг": 1.0},
                    ]
                ),
                cube_file,
            )
            repository.import_products_file_idempotent(
                run_id="quality-run",
                csv_path=cube_file,
                table_name=settings.products_table,
                project_name="unit",
                year=2025,
                month=1,
                marketplace_code="oz",
                category_key="acid",
            )
            app = create_app(settings, start_workers=False)

            with TestClient(app) as client:
                projects = client.get("/api/quality/projects")
                self.assertEqual(projects.status_code, 200)
                rows = projects.json()["projects"]
                by_name = {row["project_name"]: row for row in rows}
                self.assertEqual(by_name["unit"]["source_kind"], "cube")
                self.assertEqual(by_name["unit"]["row_count"], 2)

                report = client.get("/api/quality/report", params={"project_name": "unit"})
                self.assertEqual(report.status_code, 200)
                payload = report.json()
                self.assertEqual(payload["status"], "OK")
                self.assertEqual(payload["total_rows"], 2)
                self.assertEqual(payload["source"]["kind"], "cube")
                self.assertEqual(payload["source"]["table_name"], settings.products_table)

                missing = client.get("/api/quality/report", params={"project_name": "missing"})
                self.assertEqual(missing.status_code, 404)
                self.assertIn("не найден", missing.json()["detail"])


if __name__ == "__main__":
    unittest.main()

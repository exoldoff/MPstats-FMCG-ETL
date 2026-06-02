from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

import pandas as pd

from mpstats_app.config import AppSettings
from mpstats_app.repositories.duckdb_repository import HEAVY_CATEGORY_ROWS_LIMIT, HEAVY_SLICE_ROWS_LIMIT, DuckDbAppRepository
from mpstats_app.services.export_service import EXCEL_MAX_DATA_ROWS
from mpstats_app.utils import clean_records, clean_value


REPORT_FORMATS = {"xlsx", "csv"}
REPORT_TYPES = {
    "category_month": {
        "label": "Категория по месяцам",
        "description": "Строки, SKU, продажи, выручка, объём и цены по месяцам, категориям и маркетплейсам.",
    },
    "brand_month": {
        "label": "Бренды по месяцам",
        "description": "Та же сводка с разрезом по брендам. Удобно для pivot и динамики брендов.",
    },
    "classification_month": {
        "label": "Классификация по месяцам",
        "description": "Сводка по колонкам классификатора, если они есть в сохранённой таблице.",
    },
    "top_sku": {
        "label": "Топ SKU",
        "description": "Агрегированный топ товаров по выручке и продажам без выгрузки сырых строк.",
    },
}


@dataclass(frozen=True)
class ReportSpec:
    project_name: str
    report_type: str
    category_keys: list[str]
    period_from: str | None
    period_to: str | None
    period_from_index: int | None
    period_to_index: int | None
    export_format: str
    max_rows: int


class ReportService:
    def __init__(self, *, settings: AppSettings, repository: DuckDbAppRepository) -> None:
        self.settings = settings
        self.repository = repository

    def options(self, *, project_name: str) -> dict[str, Any]:
        project = _clean_project_name(project_name)
        payload = self.repository.report_options(table_name=self.settings.products_table, project_name=project)
        payload.update(
            {
                "project_name": project,
                "default_output_dir": str(self.default_output_dir(project)),
                "reports": [{"type": key, **value} for key, value in REPORT_TYPES.items()],
                "heavy_slice_rows_limit": HEAVY_SLICE_ROWS_LIMIT,
                "heavy_category_rows_limit": HEAVY_CATEGORY_ROWS_LIMIT,
            }
        )
        return payload

    def preview(
        self,
        *,
        project_name: str,
        report_type: str,
        category_keys: list[str],
        period_from: str | None,
        period_to: str | None,
        export_format: str,
        max_rows: int,
        limit: int,
        offset: int,
    ) -> dict[str, Any]:
        spec = self._spec(
            project_name=project_name,
            report_type=report_type,
            category_keys=category_keys,
            period_from=period_from,
            period_to=period_to,
            export_format=export_format,
            max_rows=max_rows,
        )
        total = self.repository.count_report_rows(
            table_name=self.settings.products_table,
            project_name=spec.project_name,
            report_type=spec.report_type,
            category_keys=spec.category_keys,
            period_from_index=spec.period_from_index,
            period_to_index=spec.period_to_index,
        )
        effective_limit = min(max(1, int(limit)), spec.max_rows if spec.report_type == "top_sku" else 500)
        df = self.repository.fetch_report_dataframe(
            table_name=self.settings.products_table,
            project_name=spec.project_name,
            report_type=spec.report_type,
            category_keys=spec.category_keys,
            period_from_index=spec.period_from_index,
            period_to_index=spec.period_to_index,
            limit=effective_limit,
            offset=max(0, int(offset)),
        )
        return {
            "project_name": spec.project_name,
            "report_type": spec.report_type,
            "report_label": REPORT_TYPES[spec.report_type]["label"],
            "columns": list(df.columns),
            "rows": clean_records(df.where(pd.notna(df), None).to_dict(orient="records")),
            "total": total,
            "preview_limit": effective_limit,
            "warnings": self._warnings(spec=spec, total=total),
        }

    def build(
        self,
        *,
        project_name: str,
        report_type: str,
        category_keys: list[str],
        period_from: str | None,
        period_to: str | None,
        export_format: str,
        output_dir: str | None,
        max_rows: int,
    ) -> dict[str, Any]:
        spec = self._spec(
            project_name=project_name,
            report_type=report_type,
            category_keys=category_keys,
            period_from=period_from,
            period_to=period_to,
            export_format=export_format,
            max_rows=max_rows,
        )
        total = self.repository.count_report_rows(
            table_name=self.settings.products_table,
            project_name=spec.project_name,
            report_type=spec.report_type,
            category_keys=spec.category_keys,
            period_from_index=spec.period_from_index,
            period_to_index=spec.period_to_index,
        )
        if total <= 0:
            raise ValueError("Нет строк для отчёта. Проверь проект, период и категории.")
        if spec.export_format == "xlsx" and total > EXCEL_MAX_DATA_ROWS:
            raise ValueError("Агрегированный отчёт всё ещё больше лимита Excel. Выбери CSV или сузь период/категории.")

        limit = spec.max_rows if spec.report_type == "top_sku" else None
        df = self.repository.fetch_report_dataframe(
            table_name=self.settings.products_table,
            project_name=spec.project_name,
            report_type=spec.report_type,
            category_keys=spec.category_keys,
            period_from_index=spec.period_from_index,
            period_to_index=spec.period_to_index,
            limit=limit,
            offset=0,
        )
        target_dir = self.resolve_output_dir(output_dir, spec.project_name)
        target_dir.mkdir(parents=True, exist_ok=True)
        self.repository.set_setting("report_output_dir", str(target_dir))
        target = _unique_path(target_dir / self._filename(spec, rows=len(df)))
        if spec.export_format == "csv":
            df.to_csv(target, sep=";", index=False, encoding="utf-8-sig")
        else:
            self._write_xlsx(target, df)
        self.repository.mark_reports_built(project_name=spec.project_name, category_keys=spec.category_keys)
        artifact = {
            "path": str(target),
            "filename": target.name,
            "format": spec.export_format,
            "rows": len(df),
            "report_type": spec.report_type,
            "report_label": REPORT_TYPES[spec.report_type]["label"],
        }
        return {
            "project_name": spec.project_name,
            "report_type": spec.report_type,
            "report_label": REPORT_TYPES[spec.report_type]["label"],
            "artifacts": [artifact],
            "total": len(df),
            "source_total": total,
            "output_dir": str(target_dir),
            "warnings": self._warnings(spec=spec, total=total),
        }

    def resolve_output_dir(self, output_dir: str | None, project_name: str) -> Path:
        project = _clean_project_name(project_name)
        raw = (output_dir or "").strip()
        path = Path(raw).expanduser() if raw else self.default_output_dir(project)
        if not path.is_absolute():
            path = self.settings.project_root / path
        if raw:
            path = _with_project_segment(path, project)
        return path.resolve()

    def default_output_dir(self, project_name: str) -> Path:
        return (self.settings.project_root / "data" / "projects" / _safe_segment(project_name) / "reports").resolve()

    def resolve_report_file(self, file_path: str) -> Path:
        target = Path(file_path).expanduser().resolve()
        allowed_roots = [(self.settings.project_root / "data" / "projects").resolve()]
        last_output_dir = self.repository.get_setting("report_output_dir")
        if last_output_dir:
            allowed_roots.append(Path(last_output_dir).expanduser().resolve())
        if not any(target.is_relative_to(root) for root in allowed_roots):
            raise ValueError("Можно скачать только отчёты, созданные локальным приложением.")
        if not target.exists() or not target.is_file():
            raise FileNotFoundError(f"Файл не найден: {target}")
        return target

    def _spec(
        self,
        *,
        project_name: str,
        report_type: str,
        category_keys: list[str],
        period_from: str | None,
        period_to: str | None,
        export_format: str,
        max_rows: int,
    ) -> ReportSpec:
        clean_type = report_type if report_type in REPORT_TYPES else "category_month"
        period_from_index = _period_label_to_index(period_from)
        period_to_index = _period_label_to_index(period_to)
        if period_from_index is not None and period_to_index is not None and period_from_index > period_to_index:
            raise ValueError("Начальный период отчёта больше конечного.")
        clean_format = str(export_format or "xlsx").strip().lower()
        if clean_format not in REPORT_FORMATS:
            raise ValueError("Формат отчёта должен быть xlsx или csv.")
        return ReportSpec(
            project_name=_clean_project_name(project_name),
            report_type=clean_type,
            category_keys=sorted({str(key) for key in category_keys if str(key).strip()}),
            period_from=period_from,
            period_to=period_to,
            period_from_index=period_from_index,
            period_to_index=period_to_index,
            export_format=clean_format,
            max_rows=max(100, min(int(max_rows or 5000), 100_000)),
        )

    def _filename(self, spec: ReportSpec, *, rows: int) -> str:
        stamp = datetime.now().strftime("%Y%m%d-%H%M")
        period_from = spec.period_from or "все-периоды"
        period_to = spec.period_to or "все-периоды"
        label = REPORT_TYPES[spec.report_type]["label"]
        base = f"MPStats_{spec.project_name}_{label}_{period_from}_{period_to}_{rows}стр_{stamp}"
        return f"{_safe_filename(base)}.{spec.export_format}"

    @staticmethod
    def _write_xlsx(target: Path, df: pd.DataFrame) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError as exc:
            raise ImportError("Для отчётов XLSX нужен openpyxl. Установи зависимости проекта: pip install -r requirements.txt") from exc

        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet("Отчёт")
        worksheet.freeze_panes = "A2"
        columns = [str(column) for column in df.columns]
        worksheet.append(columns)
        for row in df.itertuples(index=False, name=None):
            worksheet.append([_excel_value(value) for value in row])
        if columns:
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(df) + 1}"
        workbook.save(target)

    @staticmethod
    def _warnings(*, spec: ReportSpec, total: int) -> list[str]:
        warnings: list[str] = []
        if spec.report_type == "top_sku" and total > spec.max_rows:
            warnings.append(f"Топ SKU ограничен первыми {spec.max_rows} агрегированными строками из {total}.")
        return warnings


def _period_label_to_index(value: str | None) -> int | None:
    text = (value or "").strip()
    if not text:
        return None
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", text)
    if not match:
        raise ValueError(f"Некорректный период {value!r}. Используй формат YYYY-MM.")
    year = int(match.group(1))
    month = int(match.group(2))
    if month < 1 or month > 12:
        raise ValueError(f"Некорректный месяц в периоде {value!r}.")
    return year * 12 + month


def _clean_project_name(value: str) -> str:
    return value.strip() or "mpstats"


def _safe_segment(value: str) -> str:
    segment = re.sub(r"[^\w_.-]+", "_", value.strip(), flags=re.UNICODE)
    return segment.strip("._") or "mpstats"


def _with_project_segment(path: Path, project_name: str) -> Path:
    project_segment = _safe_segment(project_name)
    if project_segment in path.parts or project_name in path.parts:
        return path
    return path / project_segment


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[/\\:]+", "_", value)
    cleaned = re.sub(r"[^0-9A-Za-zА-Яа-яЁё._ -]+", "_", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ._")
    cleaned = re.sub(r"_+", "_", cleaned)
    return cleaned[:180] or "MPStats_report"


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(2, 1000):
        candidate = path.with_name(f"{path.stem}_v{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"Не удалось подобрать свободное имя файла для {path}")


def _excel_value(value: Any) -> Any:
    if value is pd.NA:
        return None
    cleaned = clean_value(value)
    if isinstance(cleaned, (datetime, date)):
        return cleaned
    return cleaned

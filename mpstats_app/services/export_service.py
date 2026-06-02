from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import datetime, date
import json
import math
from pathlib import Path
import re
from threading import RLock, Thread
from typing import Any
from uuid import uuid4

import pandas as pd

from mpstats_app.config import AppSettings
from mpstats_app.repositories.duckdb_repository import DuckDbAppRepository
from mpstats_app.utils import clean_records, clean_value


EXCEL_MAX_DATA_ROWS = 1_048_575
RAW_XLSX_EXPORT_ROW_LIMIT = 1_048_575
EXPORT_BATCH_SIZE = 50_000
LARGE_EXPORT_FILE_WARNING = 10
EXPORT_MATCH_TYPES = {"contains", "not_contains", "equals", "startswith", "gt", "gte", "lt", "lte"}
EXPORT_TEMPLATES_SETTING = "export_templates_json"
EXPORT_FORMATS = {"xlsx", "csv"}


@dataclass(frozen=True)
class ExportSpec:
    project_name: str
    category_keys: list[str]
    period_from: str | None
    period_to: str | None
    period_from_index: int | None
    period_to_index: int | None
    selected_columns: list[str]
    filters: list[dict[str, str]]
    excluded_row_hashes: list[str]
    sort_column: str | None
    sort_direction: str
    split_by_category: bool
    export_format: str


class ExportService:
    def __init__(self, *, settings: AppSettings, repository: DuckDbAppRepository) -> None:
        self.settings = settings
        self.repository = repository
        self._jobs: dict[str, dict[str, Any]] = {}
        self._lock = RLock()

    def options(self, *, project_name: str) -> dict[str, Any]:
        project = _clean_project_name(project_name)
        payload = self.repository.export_options(table_name=self.settings.products_table, project_name=project)
        payload.update(
            {
                "project_name": project,
                "default_output_dir": str(self.default_output_dir(project)),
                "excel_max_rows": EXCEL_MAX_DATA_ROWS,
            }
        )
        return payload

    def list_templates(self, *, project_name: str) -> dict[str, Any]:
        project = _clean_project_name(project_name)
        templates = [
            template
            for template in self._read_templates()
            if str(template.get("project_name") or "") == project
        ]
        return {"templates": sorted(templates, key=lambda item: str(item.get("updated_at") or ""), reverse=True)}

    def save_template(
        self,
        *,
        name: str,
        project_name: str,
        category_keys: list[str],
        period_from: str | None,
        period_to: str | None,
        selected_columns: list[str],
        filters: list[dict[str, str]],
        sort_column: str | None,
        sort_direction: str,
        split_by_category: bool,
        output_dir: str | None,
        export_format: str = "xlsx",
    ) -> dict[str, Any]:
        clean_name = name.strip()
        if not clean_name:
            raise ValueError("Название шаблона не заполнено.")
        spec = self._spec(
            project_name=project_name,
            category_keys=category_keys,
            period_from=period_from,
            period_to=period_to,
            selected_columns=selected_columns,
            filters=filters,
            excluded_row_hashes=[],
            sort_column=sort_column,
            sort_direction=sort_direction,
            split_by_category=split_by_category,
            export_format=export_format,
        )
        if not spec.category_keys:
            raise ValueError("Выбери хотя бы одну категорию для шаблона.")
        now = datetime.now().isoformat(timespec="seconds")
        templates = self._read_templates()
        existing = next(
            (
                template
                for template in templates
                if str(template.get("project_name") or "") == spec.project_name
                and str(template.get("name") or "").casefold() == clean_name.casefold()
            ),
            None,
        )
        template = {
            "id": str(existing.get("id")) if existing else uuid4().hex,
            "name": clean_name,
            "project_name": spec.project_name,
            "category_keys": spec.category_keys,
            "period_from": spec.period_from,
            "period_to": spec.period_to,
            "selected_columns": spec.selected_columns,
            "filters": spec.filters,
            "sort_column": spec.sort_column,
            "sort_direction": spec.sort_direction,
            "split_by_category": spec.split_by_category,
            "export_format": spec.export_format,
            "output_dir": (output_dir or "").strip() or None,
            "created_at": str(existing.get("created_at")) if existing else now,
            "updated_at": now,
        }
        templates = [item for item in templates if str(item.get("id") or "") != template["id"]]
        templates.append(template)
        self._write_templates(templates)
        return template

    def delete_template(self, *, template_id: str, project_name: str) -> dict[str, Any]:
        clean_id = template_id.strip()
        project = _clean_project_name(project_name)
        templates = self._read_templates()
        kept = [
            template
            for template in templates
            if not (str(template.get("id") or "") == clean_id and str(template.get("project_name") or "") == project)
        ]
        if len(kept) == len(templates):
            raise ValueError("Шаблон выгрузки не найден.")
        self._write_templates(kept)
        return {"template_id": clean_id, "project_name": project, "deleted": True}

    def preview(
        self,
        *,
        project_name: str,
        category_keys: list[str],
        period_from: str | None,
        period_to: str | None,
        selected_columns: list[str],
        filters: list[dict[str, str]],
        excluded_row_hashes: list[str],
        sort_column: str | None,
        sort_direction: str,
        split_by_category: bool,
        limit: int,
        offset: int,
        export_format: str = "xlsx",
    ) -> dict[str, Any]:
        spec = self._spec(
            project_name=project_name,
            category_keys=category_keys,
            period_from=period_from,
            period_to=period_to,
            selected_columns=selected_columns,
            filters=filters,
            excluded_row_hashes=excluded_row_hashes,
            sort_column=sort_column,
            sort_direction=sort_direction,
            split_by_category=split_by_category,
            export_format=export_format,
        )
        breakdown = self._breakdown(spec)
        total = self._total_from_breakdown(breakdown)
        df = self.repository.fetch_export_products_dataframe(
            table_name=self.settings.products_table,
            project_name=spec.project_name,
            output_columns=spec.selected_columns,
            category_keys=spec.category_keys,
            period_from_index=spec.period_from_index,
            period_to_index=spec.period_to_index,
            filters=spec.filters,
            excluded_row_hashes=spec.excluded_row_hashes,
            sort_column=spec.sort_column,
            sort_direction=spec.sort_direction,
            limit=limit,
            offset=offset,
            include_row_hash=True,
            default_order=bool(spec.sort_column),
        )
        estimated_files = self._estimate_file_count(spec, total, breakdown=breakdown)
        warnings = self._warnings_for_estimate(estimated_files)
        if spec.export_format == "xlsx" and total > RAW_XLSX_EXPORT_ROW_LIMIT:
            warnings.append(
                f"Raw XLSX на {total} строк не строится. Используй Данные -> Отчёты: Excel получит агрегированную таблицу."
            )
        return {
            "columns": spec.selected_columns,
            "rows": clean_records(df.where(pd.notna(df), None).to_dict(orient="records")),
            "total": total,
            "estimated_files": estimated_files,
            "export_format": spec.export_format,
            "breakdown": breakdown,
            "warnings": warnings,
        }

    def build(
        self,
        *,
        project_name: str,
        category_keys: list[str],
        period_from: str | None,
        period_to: str | None,
        selected_columns: list[str],
        filters: list[dict[str, str]],
        excluded_row_hashes: list[str],
        sort_column: str | None,
        sort_direction: str,
        split_by_category: bool,
        output_dir: str | None,
        confirm_large_export: bool,
        export_format: str = "xlsx",
        progress_callback: Any | None = None,
    ) -> dict[str, Any]:
        spec = self._spec(
            project_name=project_name,
            category_keys=category_keys,
            period_from=period_from,
            period_to=period_to,
            selected_columns=selected_columns,
            filters=filters,
            excluded_row_hashes=excluded_row_hashes,
            sort_column=sort_column,
            sort_direction=sort_direction,
            split_by_category=split_by_category,
            export_format=export_format,
        )
        breakdown = self._breakdown(spec)
        total = self._total_from_breakdown(breakdown)
        if total <= 0:
            raise ValueError("Нет строк для выгрузки. Проверь категории, период и фильтры.")
        if spec.export_format == "xlsx" and total > RAW_XLSX_EXPORT_ROW_LIMIT:
            raise ValueError(
                f"Raw XLSX на {total} строк не строится. Для больших категорий открой Данные -> Отчёты "
                "и выгрузи агрегированную таблицу для Excel."
            )

        estimated_files = self._estimate_file_count(spec, total, breakdown=breakdown)
        if estimated_files > LARGE_EXPORT_FILE_WARNING and not confirm_large_export:
            raise ValueError(
                f"Выгрузка создаст {estimated_files} файлов. Подтверди большую выгрузку и запусти экспорт ещё раз."
            )

        target_dir = self.resolve_output_dir(output_dir, spec.project_name)
        target_dir.mkdir(parents=True, exist_ok=True)
        self.repository.set_setting("export_output_dir", str(target_dir))

        artifacts: list[dict[str, Any]] = []
        completed_rows = 0
        completed_files = 0

        def report(patch: dict[str, Any]) -> None:
            if progress_callback:
                progress_callback(
                    {
                        "status": "running",
                        "total_rows": total,
                        "completed_rows": completed_rows,
                        "total_files": estimated_files,
                        "completed_files": completed_files,
                        **patch,
                    }
                )

        def row_progress(delta: int, filename: str) -> None:
            nonlocal completed_rows
            completed_rows += int(delta or 0)
            report({"current_step": f"Запись {filename}"})

        def file_progress(filename: str) -> None:
            nonlocal completed_files
            completed_files += 1
            report({"current_step": f"Готов {filename}"})

        report({"current_step": "Подготовка файлов"})
        if spec.split_by_category:
            categories = self._selected_categories_from_breakdown(breakdown)
            for category in categories:
                category_spec = replace(spec, category_keys=[str(category["category_key"])])
                category_total = int(category.get("rows_count") or 0)
                if category_total <= 0:
                    continue
                artifacts.extend(
                    self._write_parts(
                        spec=category_spec,
                        total_rows=category_total,
                        output_dir=target_dir,
                        category=category,
                        row_callback=row_progress,
                        file_callback=file_progress,
                    )
                )
        else:
            artifacts.extend(
                self._write_parts(
                    spec=spec,
                    total_rows=total,
                    output_dir=target_dir,
                    category=None,
                    row_callback=row_progress,
                    file_callback=file_progress,
                )
            )

        return {
            "artifacts": artifacts,
            "total": sum(int(item["rows"]) for item in artifacts),
            "estimated_files": estimated_files,
            "output_dir": str(target_dir),
            "split_by_category": spec.split_by_category,
            "export_format": spec.export_format,
            "breakdown": breakdown,
            "warnings": self._warnings_for_estimate(estimated_files),
        }

    def start_build_job(self, **kwargs: Any) -> dict[str, Any]:
        job_id = uuid4().hex
        job = {
            "id": job_id,
            "status": "queued",
            "progress": 0.0,
            "total_rows": 0,
            "completed_rows": 0,
            "total_files": 0,
            "completed_files": 0,
            "current_step": "В очереди",
            "error": None,
            "result": None,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        with self._lock:
            self._jobs[job_id] = job
        thread = Thread(target=self._run_build_job, args=(job_id, kwargs), daemon=True)
        thread.start()
        return self.get_build_job(job_id)

    def get_build_job(self, job_id: str) -> dict[str, Any]:
        with self._lock:
            job = self._jobs.get(job_id)
            if not job:
                raise KeyError("Задача выгрузки не найдена.")
            return {**job}

    def _run_build_job(self, job_id: str, kwargs: dict[str, Any]) -> None:
        def update(patch: dict[str, Any]) -> None:
            with self._lock:
                job = self._jobs.get(job_id)
                if not job:
                    return
                total_rows = int(patch.get("total_rows") or job.get("total_rows") or 0)
                completed_rows = int(patch.get("completed_rows") or job.get("completed_rows") or 0)
                total_files = int(patch.get("total_files") or job.get("total_files") or 0)
                completed_files = int(patch.get("completed_files") or job.get("completed_files") or 0)
                if total_rows > 0:
                    progress = min(99.0, round((completed_rows / total_rows) * 100, 1))
                elif total_files > 0:
                    progress = min(99.0, round((completed_files / total_files) * 100, 1))
                else:
                    progress = float(job.get("progress") or 0.0)
                job.update(
                    {
                        **patch,
                        "status": patch.get("status") or "running",
                        "progress": progress,
                        "updated_at": datetime.now().isoformat(timespec="seconds"),
                    }
                )

        update({"status": "running", "progress": 1.0, "current_step": "Подготовка"})
        try:
            result = self.build(**kwargs, progress_callback=update)
            with self._lock:
                job = self._jobs[job_id]
                job.update(
                    {
                        "status": "succeeded",
                        "progress": 100.0,
                        "completed_rows": int(result.get("total") or job.get("completed_rows") or 0),
                        "completed_files": len(result.get("artifacts") or []),
                        "total_files": int(result.get("estimated_files") or job.get("total_files") or 0),
                        "current_step": "Готово",
                        "result": result,
                        "updated_at": datetime.now().isoformat(timespec="seconds"),
                    }
                )
        except Exception as exc:
            with self._lock:
                job = self._jobs[job_id]
                job.update(
                    {
                        "status": "failed",
                        "error": str(exc),
                        "current_step": "Ошибка",
                        "updated_at": datetime.now().isoformat(timespec="seconds"),
                    }
                )

    def resolve_output_dir(self, output_dir: str | None, project_name: str) -> Path:
        project = _clean_project_name(project_name)
        raw = (output_dir or "").strip()
        path = Path(raw).expanduser() if raw else self.default_output_dir(project)
        if not path.is_absolute():
            path = self.settings.project_root / path
        if raw:
            path = _with_project_segment(path, project)
        return path.resolve()

    def default_output_dir(self, project_name: str) -> Path:
        return (self.settings.project_root / "data" / "projects" / _safe_segment(project_name) / "exports").resolve()

    def resolve_export_file(self, file_path: str) -> Path:
        target = Path(file_path).expanduser().resolve()
        allowed_roots = [(self.settings.project_root / "data" / "projects").resolve()]
        last_output_dir = self.repository.get_setting("export_output_dir")
        if last_output_dir:
            allowed_roots.append(Path(last_output_dir).expanduser().resolve())
        if not any(target.is_relative_to(root) for root in allowed_roots):
            raise ValueError("Можно скачать только файлы, созданные вкладкой выгрузки.")
        if not target.exists() or not target.is_file():
            raise FileNotFoundError(f"Файл не найден: {target}")
        return target

    def _spec(
        self,
        *,
        project_name: str,
        category_keys: list[str],
        period_from: str | None,
        period_to: str | None,
        selected_columns: list[str],
        filters: list[dict[str, str]],
        excluded_row_hashes: list[str],
        sort_column: str | None,
        sort_direction: str,
        split_by_category: bool,
        export_format: str,
    ) -> ExportSpec:
        visible_columns = self.repository.export_visible_columns(table_name=self.settings.products_table)
        selected = [column for column in selected_columns if column in visible_columns]
        if not selected:
            selected = visible_columns
        normalized_filters = []
        for item in filters:
            column = str(item.get("column") or "")
            value = str(item.get("value") or "").strip()
            match_type = str(item.get("match_type") or "contains")
            if column in visible_columns and value:
                clean_match_type = match_type if match_type in EXPORT_MATCH_TYPES else "contains"
                if clean_match_type in {"gt", "gte", "lt", "lte"}:
                    _parse_filter_number(value, column)
                normalized_filters.append(
                    {
                        "column": column,
                        "value": value,
                        "match_type": clean_match_type,
                    }
                )
        period_from_index = _period_label_to_index(period_from)
        period_to_index = _period_label_to_index(period_to)
        if period_from_index is not None and period_to_index is not None and period_from_index > period_to_index:
            raise ValueError("Начальный период выгрузки больше конечного.")
        clean_category_keys = sorted({str(key) for key in category_keys if str(key).strip()})
        clean_hashes = sorted({str(row_hash) for row_hash in excluded_row_hashes if str(row_hash).strip()})
        clean_sort_column = sort_column if sort_column in visible_columns else None
        return ExportSpec(
            project_name=_clean_project_name(project_name),
            category_keys=clean_category_keys,
            period_from=period_from,
            period_to=period_to,
            period_from_index=period_from_index,
            period_to_index=period_to_index,
            selected_columns=selected,
            filters=normalized_filters,
            excluded_row_hashes=clean_hashes,
            sort_column=clean_sort_column,
            sort_direction="desc" if str(sort_direction).lower() == "desc" else "asc",
            split_by_category=bool(split_by_category),
            export_format=_clean_export_format(export_format),
        )

    def _count(self, spec: ExportSpec) -> int:
        return self.repository.count_export_products(
            table_name=self.settings.products_table,
            project_name=spec.project_name,
            category_keys=spec.category_keys,
            period_from_index=spec.period_from_index,
            period_to_index=spec.period_to_index,
            filters=spec.filters,
            excluded_row_hashes=spec.excluded_row_hashes,
        )

    def _breakdown(self, spec: ExportSpec) -> list[dict[str, Any]]:
        return self.repository.export_breakdown(
            table_name=self.settings.products_table,
            project_name=spec.project_name,
            category_keys=spec.category_keys,
            period_from_index=spec.period_from_index,
            period_to_index=spec.period_to_index,
            filters=spec.filters,
            excluded_row_hashes=spec.excluded_row_hashes,
        )

    @staticmethod
    def _total_from_breakdown(breakdown: list[dict[str, Any]]) -> int:
        return sum(int(item.get("rows_count") or 0) for item in breakdown)

    def _estimate_file_count(self, spec: ExportSpec, total: int, *, breakdown: list[dict[str, Any]] | None = None) -> int:
        if total <= 0:
            return 0
        if spec.export_format == "csv":
            if not spec.split_by_category:
                return 1
            return len(self._selected_categories_from_breakdown(breakdown if breakdown is not None else self._breakdown(spec)))
        if not spec.split_by_category:
            return math.ceil(total / EXCEL_MAX_DATA_ROWS)
        categories = self._selected_categories_from_breakdown(breakdown if breakdown is not None else self._breakdown(spec))
        return sum(math.ceil(int(category.get("rows_count") or 0) / EXCEL_MAX_DATA_ROWS) for category in categories)

    @staticmethod
    def _selected_categories_from_breakdown(breakdown: list[dict[str, Any]]) -> list[dict[str, Any]]:
        categories: dict[str, dict[str, Any]] = {}
        for row in breakdown:
            category_key = str(row.get("category_key") or "")
            if not category_key:
                continue
            item = categories.setdefault(
                category_key,
                {
                    "category_key": category_key,
                    "category_name": row.get("category_name") or category_key,
                    "marketplace_code": row.get("marketplace_code"),
                    "marketplace": row.get("marketplace") or row.get("marketplace_code"),
                    "rows_count": 0,
                },
            )
            item["rows_count"] = int(item.get("rows_count") or 0) + int(row.get("rows_count") or 0)
        return sorted(
            categories.values(),
            key=lambda item: (str(item.get("category_name") or "").casefold(), str(item.get("marketplace") or "").casefold()),
        )

    def _write_parts(
        self,
        *,
        spec: ExportSpec,
        total_rows: int,
        output_dir: Path,
        category: dict[str, Any] | None,
        row_callback: Any | None = None,
        file_callback: Any | None = None,
    ) -> list[dict[str, Any]]:
        parts = max(1, math.ceil(total_rows / EXCEL_MAX_DATA_ROWS)) if spec.export_format == "xlsx" else 1
        artifacts: list[dict[str, Any]] = []
        for part_index in range(parts):
            base_offset = part_index * EXCEL_MAX_DATA_ROWS
            rows_in_part = min(EXCEL_MAX_DATA_ROWS, total_rows - base_offset) if spec.export_format == "xlsx" else total_rows
            filename = self._filename(
                spec=spec,
                total_rows=total_rows,
                category=category,
                part_index=part_index + 1,
                parts=parts,
            )
            target = _unique_path(output_dir / filename)
            if spec.export_format == "csv":
                written = self._write_csv(target=target, spec=spec, rows_in_part=rows_in_part, base_offset=base_offset, row_callback=row_callback)
            else:
                written = self._write_xlsx(target=target, spec=spec, rows_in_part=rows_in_part, base_offset=base_offset, row_callback=row_callback)
            if file_callback:
                file_callback(target.name)
            artifacts.append(
                {
                    "path": str(target),
                    "filename": target.name,
                    "format": spec.export_format,
                    "rows": written,
                    "part": part_index + 1,
                    "parts": parts,
                    "category_key": category.get("category_key") if category else None,
                    "category_name": category.get("category_name") if category else None,
                    "marketplace": category.get("marketplace") if category else None,
                }
            )
        return artifacts

    def _write_xlsx(self, *, target: Path, spec: ExportSpec, rows_in_part: int, base_offset: int, row_callback: Any | None = None) -> int:
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError as exc:
            raise ImportError("Для выгрузки XLSX нужен openpyxl. Установи зависимости проекта: pip install -r requirements.txt") from exc

        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet("Данные")
        worksheet.freeze_panes = "A2"
        worksheet.append(spec.selected_columns)
        written = 0
        while written < rows_in_part:
            batch_size = min(EXPORT_BATCH_SIZE, rows_in_part - written)
            df = self.repository.fetch_export_products_dataframe(
                table_name=self.settings.products_table,
                project_name=spec.project_name,
                output_columns=spec.selected_columns,
                category_keys=spec.category_keys,
                period_from_index=spec.period_from_index,
                period_to_index=spec.period_to_index,
                filters=spec.filters,
                excluded_row_hashes=spec.excluded_row_hashes,
                sort_column=spec.sort_column,
                sort_direction=spec.sort_direction,
                limit=batch_size,
                offset=base_offset + written,
                include_row_hash=False,
            )
            if df.empty:
                break
            for row in df.itertuples(index=False, name=None):
                worksheet.append([_excel_value(value) for value in row])
            written += len(df)
            if row_callback:
                row_callback(len(df), target.name)
        if spec.selected_columns:
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(spec.selected_columns))}{written + 1}"
        workbook.save(target)
        return written

    def _write_csv(self, *, target: Path, spec: ExportSpec, rows_in_part: int, base_offset: int, row_callback: Any | None = None) -> int:
        written = 0
        header = True
        while written < rows_in_part:
            batch_size = min(EXPORT_BATCH_SIZE, rows_in_part - written)
            df = self.repository.fetch_export_products_dataframe(
                table_name=self.settings.products_table,
                project_name=spec.project_name,
                output_columns=spec.selected_columns,
                category_keys=spec.category_keys,
                period_from_index=spec.period_from_index,
                period_to_index=spec.period_to_index,
                filters=spec.filters,
                excluded_row_hashes=spec.excluded_row_hashes,
                sort_column=spec.sort_column,
                sort_direction=spec.sort_direction,
                limit=batch_size,
                offset=base_offset + written,
                include_row_hash=False,
            )
            if df.empty:
                break
            df.to_csv(target, sep=";", index=False, mode="w" if header else "a", header=header, encoding="utf-8-sig" if header else "utf-8")
            header = False
            written += len(df)
            if row_callback:
                row_callback(len(df), target.name)
        return written

    def _filename(
        self,
        *,
        spec: ExportSpec,
        total_rows: int,
        category: dict[str, Any] | None,
        part_index: int,
        parts: int,
    ) -> str:
        stamp = datetime.now().strftime("%Y%m%d-%H%M")
        period_from = spec.period_from or "все-периоды"
        period_to = spec.period_to or "все-периоды"
        if category:
            base = (
                f"MPStats_{spec.project_name}_{category.get('category_name') or 'категория'}_"
                f"{category.get('marketplace') or category.get('marketplace_code') or 'mp'}_"
                f"{period_from}_{period_to}_{total_rows}стр_{stamp}"
            )
        else:
            base = f"MPStats_{spec.project_name}_все-категории_{period_from}_{period_to}_{total_rows}стр_{stamp}"
        if parts > 1:
            base += f"_part{part_index:02d}-of{parts:02d}"
        return f"{_safe_filename(base)}.{spec.export_format}"

    @staticmethod
    def _warnings_for_estimate(estimated_files: int) -> list[str]:
        if estimated_files <= 1:
            return []
        return [f"Выгрузка будет разбита на {estimated_files} файлов."]

    def _read_templates(self) -> list[dict[str, Any]]:
        raw = self.repository.get_setting(EXPORT_TEMPLATES_SETTING)
        if not raw:
            return []
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            return []
        templates = payload.get("templates") if isinstance(payload, dict) else payload
        if not isinstance(templates, list):
            return []
        return [item for item in templates if isinstance(item, dict)]

    def _write_templates(self, templates: list[dict[str, Any]]) -> None:
        self.repository.set_setting(
            EXPORT_TEMPLATES_SETTING,
            json.dumps({"version": 1, "templates": templates}, ensure_ascii=False),
        )


def _clean_project_name(value: str) -> str:
    return value.strip() or "mpstats"


def _safe_segment(value: str) -> str:
    segment = re.sub(r"[^\w_.-]+", "_", value.strip(), flags=re.UNICODE)
    return segment.strip("._") or "mpstats"


def _with_project_segment(path: Path, project_name: str) -> Path:
    project_segment = _safe_segment(project_name)
    if project_segment in path.parts or project_name in path.parts:
        return path
    return path / project_segment


def _parse_filter_number(value: str, column: str) -> float:
    text = value.strip().replace(",", ".")
    try:
        number = float(text)
    except ValueError as exc:
        raise ValueError(f"Фильтр {column}: для числового условия нужно число.") from exc
    if not math.isfinite(number):
        raise ValueError(f"Фильтр {column}: для числового условия нужно конечное число.")
    return number


def _clean_export_format(value: str) -> str:
    clean = str(value or "xlsx").strip().lower()
    if clean not in EXPORT_FORMATS:
        raise ValueError("Формат выгрузки должен быть xlsx или csv.")
    return clean


def _period_label_to_index(value: str | None) -> int | None:
    text = (value or "").strip()
    if not text:
        return None
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", text)
    if not match:
        raise ValueError(f"Некорректный период {value!r}. Используй формат YYYY-MM.")
    year = int(match.group(1))
    month = int(match.group(2))
    if month < 1 or month > 12:
        raise ValueError(f"Некорректный месяц в периоде {value!r}.")
    return year * 12 + month


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[/\\:]+", "_", value)
    cleaned = re.sub(r"[^0-9A-Za-zА-Яа-яЁё._ -]+", "_", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ._")
    cleaned = re.sub(r"_+", "_", cleaned)
    return cleaned[:180] or "MPStats_export"


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(2, 1000):
        candidate = path.with_name(f"{path.stem}_v{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"Не удалось подобрать свободное имя файла для {path}")


def _excel_value(value: Any) -> Any:
    if value is pd.NA:
        return None
    cleaned = clean_value(value)
    if isinstance(cleaned, (datetime, date)):
        return cleaned
    return cleaned

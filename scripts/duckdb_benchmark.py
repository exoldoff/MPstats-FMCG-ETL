from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
import json
from pathlib import Path
import shutil
import sys
import time
from typing import Callable

import duckdb
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    from duckdb_mock_data import DEFAULT_SIZES, generate_mock_csv
except ModuleNotFoundError:
    from scripts.duckdb_mock_data import DEFAULT_SIZES, generate_mock_csv
from pipeline.services.merge_service import merge_csv_files_with_duckdb
from pipeline.repositories.sql_repository import get_duckdb_connection


BENCH_SIZE_ORDER = ("small", "medium", "large")
MERGE_BENCH_SIZES = {
    "small": {"files": 3, "rows_per_file": 10_000},
    "medium": {"files": 5, "rows_per_file": 500_000},
    "large": {"files": 8, "rows_per_file": 500_000},
}
FLAT_EXPORT_SIZES = {
    "small": 10_000,
    "medium": 500_000,
    "large": 1_000_000,
}
XLSX_MAX_DATA_ROWS_WITH_HEADER = 1_048_575
MOCK_SCHEMA = """
{
    'period': 'VARCHAR',
    'date': 'DATE',
    'category': 'VARCHAR',
    'network': 'VARCHAR',
    'brand': 'VARCHAR',
    'sku': 'VARCHAR',
    'price': 'DOUBLE',
    'volume': 'DOUBLE',
    'stores_count': 'INTEGER',
    'region': 'VARCHAR'
}
"""


@dataclass(frozen=True)
class BenchResult:
    operation: str
    old_method: str
    new_method: str
    old_seconds: float | None
    new_seconds: float | None
    rows: int | None
    comment: str
    risk: str
    old_file_size_bytes: int | None = None
    new_file_size_bytes: int | None = None

    @property
    def speedup(self) -> float | None:
        if self.old_seconds is None or self.new_seconds is None or self.new_seconds <= 0:
            return None
        return self.old_seconds / self.new_seconds


@dataclass(frozen=True)
class MergeBenchResult:
    size: str
    input_files_count: int
    input_rows: int
    output_rows_old: int
    output_rows_new: int
    duplicates_removed_old: int
    duplicates_removed_new: int
    old_seconds: float
    new_seconds: float
    old_file_size_bytes: int
    new_file_size_bytes: int
    memory: str = "not_measured"

    @property
    def speedup(self) -> float | None:
        if self.new_seconds <= 0:
            return None
        return self.old_seconds / self.new_seconds


@dataclass(frozen=True)
class SettingsBenchVariant:
    name: str
    threads: int | None = None
    memory_limit: str | None = None
    temp_directory_enabled: bool = False


@dataclass(frozen=True)
class SettingsBenchResult:
    variant: str
    threads: int | None
    memory_limit: str | None
    temp_directory_enabled: bool
    load_csv_seconds: float
    report_aggregate_seconds: float
    report_csv_copy_seconds: float
    merge_csv_seconds: float | None
    loaded_rows: int
    report_rows: int
    report_csv_rows: int
    merge_output_rows: int | None
    report_csv_size_bytes: int | None
    merge_csv_size_bytes: int | None


@dataclass(frozen=True)
class FlatExportBenchResult:
    size: str
    rows: int
    old_xlsx_seconds: float
    new_xlsx_seconds: float | None
    csv_seconds: float
    old_xlsx_rows: int
    new_xlsx_rows: int | None
    csv_rows: int
    old_xlsx_size_bytes: int | None
    new_xlsx_size_bytes: int | None
    csv_size_bytes: int | None
    xlsx_opens: bool
    headers_match: bool
    numeric_cells: bool
    csv_rows_match_xlsx: bool
    row_limit_error: str | None = None

    @property
    def xlsx_speedup(self) -> float | None:
        if self.new_xlsx_seconds is None or self.new_xlsx_seconds <= 0:
            return None
        return self.old_xlsx_seconds / self.new_xlsx_seconds


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def timed(fn: Callable[[], int | None]) -> tuple[float, int | None]:
    start = time.perf_counter()
    rows = fn()
    return time.perf_counter() - start, rows


def file_size(path: Path) -> int | None:
    return path.stat().st_size if path.exists() else None


def connect(
    db_path: Path,
    *,
    threads: int | None = None,
    memory_limit: str | None = None,
    temp_directory: Path | None = None,
    use_env: bool = False,
) -> duckdb.DuckDBPyConnection:
    return get_duckdb_connection(
        db_path,
        threads=threads,
        memory_limit=memory_limit,
        temp_directory=temp_directory,
        use_env=use_env,
    )


def create_dim_tables(con: duckdb.DuckDBPyConnection) -> None:
    con.execute(
        """
        CREATE OR REPLACE TABLE dim_category AS
        SELECT *
        FROM (
            VALUES
                ('sugar', 'food'),
                ('soap', 'home'),
                ('lemon_acid', 'food'),
                ('tea', 'food'),
                ('coffee', 'food'),
                ('pasta', 'food'),
                ('cereal', 'food'),
                ('sauce', 'food')
        ) AS t(category, category_group)
        """
    )


def generate_merge_mock_csvs(workdir: Path, *, size: str, files: int, rows_per_file: int) -> list[Path]:
    input_dir = workdir / f"merge_{size}_{files}x{rows_per_file}"
    input_dir.mkdir(parents=True, exist_ok=True)
    paths = [input_dir / f"part_{index:02d}.csv" for index in range(files)]
    if all(path.exists() for path in paths):
        return paths

    fieldnames = ["SKU", "Продажи, шт", "Название", "Бренд", "Категория"]
    for file_index, path in enumerate(paths):
        with path.open("w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=fieldnames, delimiter=";")
            writer.writeheader()
            for row_index in range(rows_per_file):
                shared = row_index % 4 == 0
                sku = f"shared-{row_index:08d}" if shared else f"file{file_index:02d}-{row_index:08d}"
                sales: float | int
                if row_index % 997 == 0:
                    sales = 0
                elif row_index % 991 == 0:
                    sales = 50_000
                else:
                    sales = 1 + row_index % 1000
                writer.writerow(
                    {
                        "SKU": sku,
                        "Продажи, шт": sales,
                        "Название": f"product {row_index:08d}",
                        "Бренд": f"Brand {row_index % 80:02d}",
                        "Категория": f"Category {row_index % 12:02d}",
                    }
                )
    return paths


def old_merge_csv_export(input_paths: list[Path], output_path: Path) -> dict[str, int]:
    frames = [pd.read_csv(path, sep=";", encoding="utf-8-sig", low_memory=False) for path in input_paths]
    input_rows = sum(len(frame) for frame in frames)
    merged = pd.concat(frames, ignore_index=True)
    merged["Продажи, шт"] = (
        merged["Продажи, шт"].astype(str).str.replace(" ", "", regex=False).str.replace("\u00a0", "", regex=False).str.replace(",", ".", regex=False)
    )
    merged["Продажи, шт"] = pd.to_numeric(merged["Продажи, шт"], errors="coerce").fillna(0)
    filtered = merged[(merged["Продажи, шт"] > 0) & (merged["Продажи, шт"] < 40_000)].copy()
    deduped = filtered.drop_duplicates()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    deduped.to_csv(output_path, sep=";", index=False, encoding="utf-8-sig")
    return {
        "input_rows": input_rows,
        "filtered_rows": len(filtered),
        "output_rows": len(deduped),
        "duplicates_removed": len(filtered) - len(deduped),
    }


def run_merge_benchmark(*, workdir: Path, size: str) -> MergeBenchResult:
    config = MERGE_BENCH_SIZES[size]
    input_paths = generate_merge_mock_csvs(
        workdir,
        size=size,
        files=int(config["files"]),
        rows_per_file=int(config["rows_per_file"]),
    )
    old_output = workdir / f"merge_old_{size}.csv"
    new_output = workdir / f"merge_new_{size}.csv"

    old_stats: dict[str, int] = {}

    def old_merge() -> int:
        nonlocal old_stats
        old_stats = old_merge_csv_export(input_paths, old_output)
        return old_stats["output_rows"]

    old_seconds, _ = timed(old_merge)

    new_result_holder: dict[str, object] = {}

    def new_merge() -> int:
        result = merge_csv_files_with_duckdb(input_paths, new_output)
        new_result_holder["result"] = result
        return result.rows_out

    new_seconds, _ = timed(new_merge)
    new_result = new_result_holder["result"]
    return MergeBenchResult(
        size=size,
        input_files_count=len(input_paths),
        input_rows=int(old_stats["input_rows"]),
        output_rows_old=int(old_stats["output_rows"]),
        output_rows_new=int(new_result.rows_out),
        duplicates_removed_old=int(old_stats["duplicates_removed"]),
        duplicates_removed_new=int(new_result.duplicates_removed),
        old_seconds=old_seconds,
        new_seconds=new_seconds,
        old_file_size_bytes=file_size(old_output) or 0,
        new_file_size_bytes=file_size(new_output) or 0,
    )


def duckdb_csv_scan(csv_path: Path) -> str:
    return (
        "read_csv("
        f"{sql_literal(str(csv_path))}, "
        "delim=';', header=true, dateformat='%Y-%m-%d', "
        f"columns={MOCK_SCHEMA}"
        ")"
    )


def stage_select(csv_path: Path) -> str:
    scan = duckdb_csv_scan(csv_path)
    return f"""
        SELECT
            period,
            date,
            category,
            network,
            brand,
            sku,
            price,
            volume,
            stores_count,
            region,
            hash(period, date, category, network, brand, sku, price, volume, stores_count, region)::VARCHAR AS row_hash
        FROM {scan}
        WHERE price > 0 AND volume > 0
    """


GROUP_QUERY = """
    SELECT
        period,
        category,
        network,
        COUNT(*) AS rows_count,
        SUM(volume) AS total_volume,
        SUM(price * volume) AS revenue,
        AVG(price) AS avg_price
    FROM products_new
    WHERE period BETWEEN '2024-06' AND '2025-06'
      AND category IN ('sugar', 'tea', 'coffee')
      AND network IN ('Ozon', 'Wildberries')
    GROUP BY period, category, network
    ORDER BY period, category, network
"""

REPORT_CSV_QUERY = """
    SELECT
        period,
        category,
        network,
        brand,
        sku,
        COUNT(*) AS rows_count,
        SUM(volume) AS total_volume,
        SUM(price * volume) AS revenue,
        AVG(price) AS avg_price
    FROM products_new
    WHERE period BETWEEN '2024-01' AND '2025-12'
    GROUP BY period, category, network, brand, sku
    ORDER BY revenue DESC NULLS LAST, period, category, network, brand, sku
"""

FLAT_EXPORT_QUERY = """
    SELECT
        date,
        period,
        category,
        network,
        brand,
        sku,
        price,
        volume,
        stores_count,
        price * volume AS revenue
    FROM products_new
    ORDER BY sku
"""

JOIN_QUERY = """
    SELECT
        p.period,
        d.category_group,
        p.network,
        COUNT(*) AS rows_count,
        SUM(p.volume) AS total_volume
    FROM products_new p
    JOIN dim_category d USING (category)
    WHERE p.period >= '2024-06'
    GROUP BY p.period, d.category_group, p.network
"""


def run_benchmark(
    *,
    workdir: Path,
    size: str,
    rows: int,
    threads: int | None,
    memory_limit: str | None,
    skip_excel: bool,
) -> tuple[list[BenchResult], dict[str, object]]:
    workdir.mkdir(parents=True, exist_ok=True)
    csv_path = workdir / f"mock_{size}_{rows}.csv"
    if not csv_path.exists():
        generate_mock_csv(csv_path, rows=rows)

    old_db = workdir / f"old_{size}.duckdb"
    new_db = workdir / f"new_{size}.duckdb"
    for path in (old_db, new_db):
        if path.exists():
            path.unlink()

    results: list[BenchResult] = []

    with connect(old_db, threads=threads, memory_limit=memory_limit, temp_directory=workdir / "tmp_old") as con:
        def old_load() -> int:
            df = pd.read_csv(csv_path, sep=";", low_memory=False)
            con.register("mock_df", df)
            con.execute("CREATE OR REPLACE TABLE products_old AS SELECT * FROM mock_df")
            con.unregister("mock_df")
            return int(con.execute("SELECT COUNT(*) FROM products_old").fetchone()[0])

        old_load_seconds, old_rows = timed(old_load)

    with connect(new_db, threads=threads, memory_limit=memory_limit, temp_directory=workdir / "tmp_new") as con:
        create_dim_tables(con)

        def new_load() -> int:
            con.execute(f"CREATE OR REPLACE TABLE products_new AS {stage_select(csv_path)}")
            return int(con.execute("SELECT COUNT(*) FROM products_new").fetchone()[0])

        new_load_seconds, new_rows = timed(new_load)
        results.append(
            BenchResult(
                "load_csv_to_duckdb",
                "pandas.read_csv -> register -> CTAS",
                "DuckDB read_csv -> CTAS with explicit schema",
                old_load_seconds,
                new_load_seconds,
                new_rows,
                "DuckDB filters zero-volume rows during load.",
                "CSV dialect is fixed to semicolon/UTF-8 for this benchmark.",
            )
        )

        def idempotent_rerun() -> int:
            con.execute("CREATE OR REPLACE TEMP TABLE stage_products AS " + stage_select(csv_path))
            before = int(con.execute("SELECT COUNT(*) FROM products_new").fetchone()[0])
            con.execute(
                """
                INSERT INTO products_new
                SELECT s.*
                FROM stage_products s
                WHERE NOT EXISTS (
                    SELECT 1 FROM products_new p WHERE p.row_hash = s.row_hash
                )
                """
            )
            after = int(con.execute("SELECT COUNT(*) FROM products_new").fetchone()[0])
            return after - before

        rerun_seconds, inserted_again = timed(idempotent_rerun)
        results.append(
            BenchResult(
                "idempotent_rerun",
                "append can duplicate rows",
                "stage table + anti-join by row_hash",
                None,
                rerun_seconds,
                inserted_again,
                "Expected inserted_again=0.",
                "row_hash key must match production business key policy.",
            )
        )

        def group_by() -> int:
            return len(con.execute(GROUP_QUERY).fetchall())

        group_seconds, group_rows = timed(group_by)
        results.append(
            BenchResult(
                "group_by_report",
                "fetch raw rows to pandas, then aggregate",
                "DuckDB GROUP BY SQL",
                None,
                group_seconds,
                group_rows,
                "Heavy aggregation stays inside DuckDB.",
                "No pandas baseline is run here because it scales poorly by design.",
            )
        )

        def join_query() -> int:
            return len(con.execute(JOIN_QUERY).fetchall())

        join_seconds, join_rows = timed(join_query)
        results.append(
            BenchResult(
                "join_with_dictionary",
                "pandas merge after fetching raw data",
                "DuckDB SQL join",
                None,
                join_seconds,
                join_rows,
                "Dictionary join stays inside DuckDB.",
                "Representative mock dictionary only.",
            )
        )

        def filtered_count() -> int:
            return int(
                con.execute(
                    """
                    SELECT COUNT(*)
                    FROM products_new
                    WHERE period BETWEEN '2024-06' AND '2025-06'
                      AND category = 'sugar'
                      AND network = 'Ozon'
                    """
                ).fetchone()[0]
            )

        filter_seconds, filter_rows = timed(filtered_count)
        results.append(
            BenchResult(
                "filter_period_category_network",
                "pandas boolean mask after fetch",
                "DuckDB WHERE filter",
                None,
                filter_seconds,
                filter_rows,
                "Filters are pushed before result materialization.",
                "No indexes: speed depends on DuckDB scan and table ordering.",
            )
        )

        old_csv = workdir / f"report_old_{size}.csv"
        new_csv = workdir / f"report_new_{size}.csv"

        def old_csv_export() -> int:
            df = con.execute(REPORT_CSV_QUERY).fetchdf()
            df.to_csv(old_csv, sep=";", index=False, encoding="utf-8-sig")
            return len(df)

        old_csv_seconds, old_csv_rows = timed(old_csv_export)

        def new_csv_export() -> int:
            row = con.execute(f"COPY ({REPORT_CSV_QUERY}) TO {sql_literal(str(new_csv))} WITH (FORMAT csv, DELIMITER ';', HEADER true)").fetchone()
            return int(row[0]) if row else 0

        new_csv_seconds, new_csv_rows = timed(new_csv_export)
        results.append(
            BenchResult(
                "report_csv_export",
                "fetchdf -> df.to_csv",
                "COPY (SELECT ...) TO CSV",
                old_csv_seconds,
                new_csv_seconds,
                new_csv_rows,
                "Direct report CSV export avoids pandas materialization.",
                "COPY output has simpler formatting controls.",
                old_file_size_bytes=file_size(old_csv),
                new_file_size_bytes=file_size(new_csv),
            )
        )

        if not skip_excel:
            old_xlsx = workdir / f"report_old_{size}.xlsx"
            new_xlsx = workdir / f"report_new_{size}.xlsx"

            def old_xlsx_export() -> int:
                df = con.execute(GROUP_QUERY).fetchdf()
                df.to_excel(old_xlsx, index=False)
                return len(df)

            old_xlsx_seconds, old_xlsx_rows = timed(old_xlsx_export)
            try:
                def new_xlsx_export() -> int:
                    con.execute("INSTALL excel")
                    con.execute("LOAD excel")
                    con.execute(f"COPY ({GROUP_QUERY}) TO {sql_literal(str(new_xlsx))} WITH (FORMAT xlsx, HEADER true)")
                    return int(new_xlsx.exists() and new_xlsx.stat().st_size > 0)

                new_xlsx_seconds, _ = timed(new_xlsx_export)
                xlsx_comment = "DuckDB Excel extension works for plain XLSX."
                xlsx_risk = "Use openpyxl for styled/multi-sheet workbooks."
            except Exception as exc:
                new_xlsx_seconds = None
                xlsx_comment = f"DuckDB XLSX export unavailable: {exc}"
                xlsx_risk = "Extension install/load may be unavailable offline."
            results.append(
                BenchResult(
                    "export_xlsx",
                    "fetchdf -> pandas.to_excel",
                    "COPY (SELECT ...) TO XLSX via DuckDB excel extension",
                    old_xlsx_seconds,
                    new_xlsx_seconds,
                    old_xlsx_rows,
                    xlsx_comment,
                    xlsx_risk,
                )
            )

        quality = {
            "source_csv": str(csv_path),
            "source_size_bytes": csv_path.stat().st_size,
            "requested_rows": rows,
            "loaded_rows": new_rows,
            "duplicate_row_hashes": int(
                con.execute(
                    """
                    SELECT COALESCE(SUM(cnt - 1), 0)
                    FROM (
                        SELECT row_hash, COUNT(*) AS cnt
                        FROM products_new
                        GROUP BY row_hash
                        HAVING COUNT(*) > 1
                    )
                    """
                ).fetchone()[0]
            ),
            "null_required_rows": int(
                con.execute(
                    """
                    SELECT COUNT(*)
                    FROM products_new
                    WHERE period IS NULL OR date IS NULL OR category IS NULL OR network IS NULL OR sku IS NULL
                    """
                ).fetchone()[0]
            ),
            "date_min": str(con.execute("SELECT MIN(date) FROM products_new").fetchone()[0]),
            "date_max": str(con.execute("SELECT MAX(date) FROM products_new").fetchone()[0]),
            "idempotent_inserted_again": inserted_again,
            "csv_export_exists": new_csv.exists() and new_csv.stat().st_size > 0,
            "report_csv_rows_old": old_csv_rows,
            "report_csv_rows_new": new_csv_rows,
            "report_csv_old_size_bytes": file_size(old_csv),
            "report_csv_new_size_bytes": file_size(new_csv),
            "memory": "not_measured",
        }

    shutil.rmtree(workdir / "tmp_old", ignore_errors=True)
    shutil.rmtree(workdir / "tmp_new", ignore_errors=True)
    return results, quality


def settings_variants() -> list[SettingsBenchVariant]:
    return [
        SettingsBenchVariant("default"),
        SettingsBenchVariant("threads=1", threads=1),
        SettingsBenchVariant("threads=4", threads=4),
        SettingsBenchVariant("memory_limit=2GB", memory_limit="2GB"),
        SettingsBenchVariant("memory_limit=4GB", memory_limit="4GB"),
        SettingsBenchVariant("temp_directory", temp_directory_enabled=True),
    ]


def safe_variant_name(name: str) -> str:
    return (
        name.lower()
        .replace("=", "_")
        .replace(" ", "_")
        .replace("/", "_")
        .replace("-", "_")
    )


def run_settings_benchmark(
    *,
    workdir: Path,
    size: str,
    rows: int,
    merge_size: str,
    include_merge: bool,
) -> list[SettingsBenchResult]:
    workdir.mkdir(parents=True, exist_ok=True)
    csv_path = workdir / f"mock_settings_{size}_{rows}.csv"
    if not csv_path.exists():
        generate_mock_csv(csv_path, rows=rows)

    merge_input_paths: list[Path] = []
    if include_merge:
        merge_config = MERGE_BENCH_SIZES[merge_size]
        merge_input_paths = generate_merge_mock_csvs(
            workdir,
            size=merge_size,
            files=int(merge_config["files"]),
            rows_per_file=int(merge_config["rows_per_file"]),
        )

    results: list[SettingsBenchResult] = []
    for variant in settings_variants():
        variant_slug = safe_variant_name(variant.name)
        db_path = workdir / f"settings_{variant_slug}.duckdb"
        if db_path.exists():
            db_path.unlink()
        report_csv = workdir / f"settings_report_{variant_slug}.csv"
        if report_csv.exists():
            report_csv.unlink()
        temp_directory = workdir / f"settings_tmp_{variant_slug}" if variant.temp_directory_enabled else None

        with connect(
            db_path,
            threads=variant.threads,
            memory_limit=variant.memory_limit,
            temp_directory=temp_directory,
            use_env=False,
        ) as con:
            create_dim_tables(con)

            def load_csv() -> int:
                con.execute(f"CREATE OR REPLACE TABLE products_new AS {stage_select(csv_path)}")
                return int(con.execute("SELECT COUNT(*) FROM products_new").fetchone()[0])

            load_seconds, loaded_rows = timed(load_csv)

            def report_aggregate() -> int:
                return len(con.execute(GROUP_QUERY).fetchall())

            aggregate_seconds, report_rows = timed(report_aggregate)

            def report_csv_copy() -> int:
                row = con.execute(
                    f"COPY ({REPORT_CSV_QUERY}) TO {sql_literal(str(report_csv))} "
                    "WITH (FORMAT csv, DELIMITER ';', HEADER true)"
                ).fetchone()
                return int(row[0]) if row else 0

            copy_seconds, report_csv_rows = timed(report_csv_copy)

        merge_seconds: float | None = None
        merge_output_rows: int | None = None
        merge_csv_size_bytes: int | None = None
        if include_merge:
            merge_output = workdir / f"settings_merge_{variant_slug}_{merge_size}.csv"
            if merge_output.exists():
                merge_output.unlink()
            merge_temp_directory = temp_directory / "merge" if temp_directory is not None else None
            merge_result = merge_csv_files_with_duckdb(
                merge_input_paths,
                merge_output,
                duckdb_threads=variant.threads,
                duckdb_memory_limit=variant.memory_limit,
                duckdb_temp_directory=merge_temp_directory,
                use_env_settings=False,
            )
            merge_seconds = merge_result.duration_seconds
            merge_output_rows = merge_result.rows_out
            merge_csv_size_bytes = merge_result.file_size_bytes

        results.append(
            SettingsBenchResult(
                variant=variant.name,
                threads=variant.threads,
                memory_limit=variant.memory_limit,
                temp_directory_enabled=variant.temp_directory_enabled,
                load_csv_seconds=load_seconds,
                report_aggregate_seconds=aggregate_seconds,
                report_csv_copy_seconds=copy_seconds,
                merge_csv_seconds=merge_seconds,
                loaded_rows=int(loaded_rows or 0),
                report_rows=int(report_rows or 0),
                report_csv_rows=int(report_csv_rows or 0),
                merge_output_rows=merge_output_rows,
                report_csv_size_bytes=file_size(report_csv),
                merge_csv_size_bytes=merge_csv_size_bytes,
            )
        )
        if temp_directory is not None:
            shutil.rmtree(temp_directory, ignore_errors=True)

    return results


def old_openpyxl_flat_export(con: duckdb.DuckDBPyConnection, query: str, output_path: Path, *, batch_size: int = 50_000) -> int:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required for old flat XLSX benchmark") from exc

    metadata = con.execute(f"SELECT * FROM ({query}) AS flat_source LIMIT 0")
    columns = [str(item[0]) for item in (metadata.description or [])]
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Data")
    worksheet.append(columns)

    written = 0
    while True:
        rows = con.execute(
            f"SELECT * FROM ({query}) AS flat_source LIMIT ? OFFSET ?",
            [int(batch_size), int(written)],
        ).fetchall()
        if not rows:
            break
        for row in rows:
            worksheet.append(list(row))
        written += len(rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return written


def duckdb_xlsx_flat_export(con: duckdb.DuckDBPyConnection, query: str, output_path: Path) -> int:
    count_row = con.execute(f"SELECT COUNT(*) FROM ({query}) AS flat_source").fetchone()
    rows = int(count_row[0]) if count_row else 0
    if rows > XLSX_MAX_DATA_ROWS_WITH_HEADER:
        raise ValueError(f"XLSX row limit exceeded ({rows} rows), use CSV.")
    con.execute("INSTALL excel")
    con.execute("LOAD excel")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    row = con.execute(
        f"COPY ({query}) TO {sql_literal(str(output_path))} WITH (FORMAT xlsx, HEADER true, SHEET 'Data')"
    ).fetchone()
    return int(row[0]) if row and row[0] is not None else rows


def duckdb_csv_flat_export(con: duckdb.DuckDBPyConnection, query: str, output_path: Path) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    row = con.execute(
        f"COPY ({query}) TO {sql_literal(str(output_path))} WITH (FORMAT csv, DELIMITER ';', HEADER true)"
    ).fetchone()
    return int(row[0]) if row and row[0] is not None else 0


def validate_flat_xlsx(path: Path, *, expected_headers: list[str]) -> tuple[bool, bool, bool]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return False, False, False
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    first_row = list(next(rows))
    xlsx_opens = True
    headers_match = headers == expected_headers
    numeric_cells = isinstance(first_row[6], (int, float)) and isinstance(first_row[7], (int, float)) and isinstance(first_row[8], int) and isinstance(first_row[9], (int, float))
    workbook.close()
    return xlsx_opens, headers_match, numeric_cells


def run_flat_export_benchmark(*, workdir: Path, size: str, rows: int) -> FlatExportBenchResult:
    workdir.mkdir(parents=True, exist_ok=True)
    csv_path = workdir / f"mock_flat_{size}_{rows}.csv"
    if not csv_path.exists():
        generate_mock_csv(csv_path, rows=rows)

    db_path = workdir / f"flat_export_{size}.duckdb"
    if db_path.exists():
        db_path.unlink()
    old_xlsx = workdir / f"flat_old_{size}.xlsx"
    new_xlsx = workdir / f"flat_new_{size}.xlsx"
    csv_output = workdir / f"flat_new_{size}.csv"
    for path in (old_xlsx, new_xlsx, csv_output):
        if path.exists():
            path.unlink()

    with connect(db_path, use_env=False) as con:
        con.execute(f"CREATE OR REPLACE TABLE products_new AS {stage_select(csv_path)}")
        expected_rows = int(con.execute("SELECT COUNT(*) FROM products_new").fetchone()[0])

        old_xlsx_seconds, old_xlsx_rows = timed(lambda: old_openpyxl_flat_export(con, FLAT_EXPORT_QUERY, old_xlsx))

        new_xlsx_seconds: float | None
        new_xlsx_rows: int | None
        try:
            new_xlsx_seconds, new_xlsx_rows = timed(lambda: duckdb_xlsx_flat_export(con, FLAT_EXPORT_QUERY, new_xlsx))
        except Exception:
            new_xlsx_seconds = None
            new_xlsx_rows = None

        csv_seconds, csv_rows = timed(lambda: duckdb_csv_flat_export(con, FLAT_EXPORT_QUERY, csv_output))

        row_limit_error = None
        try:
            duckdb_xlsx_flat_export(con, "SELECT range AS id FROM range(1048577)", workdir / "flat_row_limit_should_not_exist.xlsx")
        except ValueError as exc:
            row_limit_error = str(exc)

    headers = ["date", "period", "category", "network", "brand", "sku", "price", "volume", "stores_count", "revenue"]
    if new_xlsx.exists():
        xlsx_opens, headers_match, numeric_cells = validate_flat_xlsx(new_xlsx, expected_headers=headers)
    else:
        xlsx_opens, headers_match, numeric_cells = False, False, False

    return FlatExportBenchResult(
        size=size,
        rows=expected_rows,
        old_xlsx_seconds=old_xlsx_seconds,
        new_xlsx_seconds=new_xlsx_seconds,
        csv_seconds=csv_seconds,
        old_xlsx_rows=int(old_xlsx_rows or 0),
        new_xlsx_rows=None if new_xlsx_rows is None else int(new_xlsx_rows),
        csv_rows=int(csv_rows or 0),
        old_xlsx_size_bytes=file_size(old_xlsx),
        new_xlsx_size_bytes=file_size(new_xlsx),
        csv_size_bytes=file_size(csv_output),
        xlsx_opens=xlsx_opens,
        headers_match=headers_match,
        numeric_cells=numeric_cells,
        csv_rows_match_xlsx=bool(new_xlsx_rows is not None and int(csv_rows or 0) == int(new_xlsx_rows)),
        row_limit_error=row_limit_error,
    )


def markdown_table(results: list[BenchResult]) -> str:
    lines = [
        "| operation | old method | new method | before, s | after, s | speedup | rows | old bytes | new bytes | comment | risk |",
        "| --- | --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | --- | --- |",
    ]
    for item in results:
        before = "-" if item.old_seconds is None else f"{item.old_seconds:.4f}"
        after = "-" if item.new_seconds is None else f"{item.new_seconds:.4f}"
        speedup = "-" if item.speedup is None else f"{item.speedup:.2f}x"
        rows = "-" if item.rows is None else str(item.rows)
        old_bytes = "-" if item.old_file_size_bytes is None else str(item.old_file_size_bytes)
        new_bytes = "-" if item.new_file_size_bytes is None else str(item.new_file_size_bytes)
        lines.append(
            "| "
            + " | ".join(
                [
                    item.operation,
                    item.old_method,
                    item.new_method,
                    before,
                    after,
                    speedup,
                    rows,
                    old_bytes,
                    new_bytes,
                    item.comment.replace("|", "/"),
                    item.risk.replace("|", "/"),
                ]
            )
            + " |"
        )
    return "\n".join(lines)


def merge_markdown_table(results: list[MergeBenchResult]) -> str:
    lines = [
        "| size | input files | input rows | old output rows | new output rows | old dup removed | new dup removed | before, s | after, s | speedup | old bytes | new bytes | memory |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | --- |",
    ]
    for item in results:
        speedup = "-" if item.speedup is None else f"{item.speedup:.2f}x"
        lines.append(
            "| "
            + " | ".join(
                [
                    item.size,
                    str(item.input_files_count),
                    str(item.input_rows),
                    str(item.output_rows_old),
                    str(item.output_rows_new),
                    str(item.duplicates_removed_old),
                    str(item.duplicates_removed_new),
                    f"{item.old_seconds:.4f}",
                    f"{item.new_seconds:.4f}",
                    speedup,
                    str(item.old_file_size_bytes),
                    str(item.new_file_size_bytes),
                    item.memory,
                ]
            )
            + " |"
        )
    return "\n".join(lines)


def settings_markdown_table(results: list[SettingsBenchResult]) -> str:
    default_total = None
    for item in results:
        if item.variant == "default":
            default_total = item.load_csv_seconds + item.report_aggregate_seconds + item.report_csv_copy_seconds + (item.merge_csv_seconds or 0)
            break
    lines = [
        "| variant | threads | memory_limit | temp dir | load CSV, s | report aggregate, s | report CSV COPY, s | merge CSV, s | total, s | speedup vs default | rows loaded | report rows | merge rows |",
        "| --- | ---: | --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for item in results:
        total = item.load_csv_seconds + item.report_aggregate_seconds + item.report_csv_copy_seconds + (item.merge_csv_seconds or 0)
        speedup = "-" if not default_total or total <= 0 else f"{default_total / total:.2f}x"
        merge_seconds = "-" if item.merge_csv_seconds is None else f"{item.merge_csv_seconds:.4f}"
        merge_rows = "-" if item.merge_output_rows is None else str(item.merge_output_rows)
        lines.append(
            "| "
            + " | ".join(
                [
                    item.variant,
                    "-" if item.threads is None else str(item.threads),
                    "-" if item.memory_limit is None else item.memory_limit,
                    "yes" if item.temp_directory_enabled else "no",
                    f"{item.load_csv_seconds:.4f}",
                    f"{item.report_aggregate_seconds:.4f}",
                    f"{item.report_csv_copy_seconds:.4f}",
                    merge_seconds,
                    f"{total:.4f}",
                    speedup,
                    str(item.loaded_rows),
                    str(item.report_csv_rows),
                    merge_rows,
                ]
            )
            + " |"
        )
    return "\n".join(lines)


def flat_export_markdown_table(results: list[FlatExportBenchResult]) -> str:
    lines = [
        "| size | rows | old XLSX openpyxl, s | new XLSX DuckDB COPY, s | XLSX speedup | CSV COPY, s | old XLSX rows | new XLSX rows | CSV rows | old XLSX bytes | new XLSX bytes | CSV bytes | xlsx opens | headers match | numeric cells | CSV rows match XLSX | row-limit error |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | --- | --- | --- | --- | --- |",
    ]
    for item in results:
        new_xlsx_seconds = "-" if item.new_xlsx_seconds is None else f"{item.new_xlsx_seconds:.4f}"
        speedup = "-" if item.xlsx_speedup is None else f"{item.xlsx_speedup:.2f}x"
        new_xlsx_rows = "-" if item.new_xlsx_rows is None else str(item.new_xlsx_rows)
        lines.append(
            "| "
            + " | ".join(
                [
                    item.size,
                    str(item.rows),
                    f"{item.old_xlsx_seconds:.4f}",
                    new_xlsx_seconds,
                    speedup,
                    f"{item.csv_seconds:.4f}",
                    str(item.old_xlsx_rows),
                    new_xlsx_rows,
                    str(item.csv_rows),
                    "-" if item.old_xlsx_size_bytes is None else str(item.old_xlsx_size_bytes),
                    "-" if item.new_xlsx_size_bytes is None else str(item.new_xlsx_size_bytes),
                    "-" if item.csv_size_bytes is None else str(item.csv_size_bytes),
                    str(item.xlsx_opens).lower(),
                    str(item.headers_match).lower(),
                    str(item.numeric_cells).lower(),
                    str(item.csv_rows_match_xlsx).lower(),
                    "-" if item.row_limit_error is None else item.row_limit_error.replace("|", "/"),
                ]
            )
            + " |"
        )
    return "\n".join(lines)


def merge_sizes_for_args(*, size: str, all_sizes: bool, include_large_merge: bool) -> list[str]:
    if all_sizes:
        sizes = ["small", "medium"]
        if include_large_merge:
            sizes.append("large")
        return sizes
    if size == "large" and not include_large_merge:
        raise ValueError("Merge benchmark size=large запускается только с --include-large-merge.")
    return [size]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Benchmark old pandas-heavy paths against DuckDB SQL paths on mock data.")
    parser.add_argument("--workdir", type=Path, default=Path("data/duckdb_benchmark"), help="Directory for generated data and benchmark outputs.")
    parser.add_argument("--size", choices=sorted(DEFAULT_SIZES), default="small", help="Named mock dataset size.")
    parser.add_argument("--rows", type=int, default=None, help="Override row count.")
    parser.add_argument("--threads", type=int, default=None, help="DuckDB threads setting.")
    parser.add_argument("--memory-limit", default=None, help="DuckDB memory_limit setting, e.g. 4GB.")
    parser.add_argument("--skip-excel", action="store_true", help="Skip XLSX export benchmarks.")
    parser.add_argument("--all-sizes", action="store_true", help="Run small, medium and large sizes in one pass.")
    parser.add_argument("--merge-only", action="store_true", help="Run only merge CSV benchmark.")
    parser.add_argument("--include-large-merge", action="store_true", help="Allow the large merge benchmark.")
    parser.add_argument("--settings-only", action="store_true", help="Run DuckDB settings benchmark: default, threads, memory_limit and temp_directory.")
    parser.add_argument("--settings-merge-size", choices=sorted(MERGE_BENCH_SIZES), default="medium", help="Merge dataset size for --settings-only.")
    parser.add_argument("--settings-skip-merge", action="store_true", help="Skip merge operation inside --settings-only.")
    parser.add_argument("--flat-export-only", action="store_true", help="Run flat CSV/XLSX export benchmark.")
    parser.add_argument("--include-large-flat-export", action="store_true", help="Allow the large 1,000,000-row flat export benchmark.")
    return parser.parse_args()


def write_outputs(*, workdir: Path, size: str, rows: int, results: list[BenchResult], quality: dict[str, object]) -> str:
    payload = {
        "size": size,
        "rows": rows,
        "quality": quality,
        "results": [
            {
                **item.__dict__,
                "speedup": item.speedup,
            }
            for item in results
        ],
    }
    workdir.mkdir(parents=True, exist_ok=True)
    (workdir / f"benchmark_{size}.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    report = markdown_table(results)
    (workdir / f"benchmark_{size}.md").write_text(report + "\n", encoding="utf-8")
    return report


def write_settings_outputs(*, workdir: Path, size: str, rows: int, merge_size: str, results: list[SettingsBenchResult]) -> str:
    payload = {
        "benchmark": "duckdb_settings",
        "size": size,
        "rows": rows,
        "merge_size": merge_size,
        "results": [item.__dict__ for item in results],
    }
    workdir.mkdir(parents=True, exist_ok=True)
    (workdir / "settings_benchmark.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    report = settings_markdown_table(results)
    (workdir / "settings_benchmark.md").write_text(report + "\n", encoding="utf-8")
    return report


def write_flat_export_outputs(*, workdir: Path, results: list[FlatExportBenchResult]) -> str:
    payload = {
        "benchmark": "flat_export",
        "results": [
            {
                **item.__dict__,
                "xlsx_speedup": item.xlsx_speedup,
            }
            for item in results
        ],
    }
    workdir.mkdir(parents=True, exist_ok=True)
    (workdir / "flat_export_benchmark.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    report = flat_export_markdown_table(results)
    (workdir / "flat_export_benchmark.md").write_text(report + "\n", encoding="utf-8")
    return report


def main() -> None:
    args = parse_args()
    if args.all_sizes and args.rows is not None:
        raise ValueError("--rows нельзя использовать вместе с --all-sizes.")
    if args.flat_export_only:
        if args.rows is not None and args.all_sizes:
            raise ValueError("--rows нельзя использовать вместе с --flat-export-only --all-sizes.")
        if args.all_sizes:
            sizes = ["small", "medium"]
            if args.include_large_flat_export:
                sizes.append("large")
        else:
            if args.size == "large" and not args.include_large_flat_export:
                raise ValueError("Flat export benchmark size=large запускается только с --include-large-flat-export.")
            sizes = [args.size]
        results = []
        for size in sizes:
            rows = int(args.rows if args.rows is not None else FLAT_EXPORT_SIZES[size])
            results.append(run_flat_export_benchmark(workdir=args.workdir, size=size, rows=rows))
        report = write_flat_export_outputs(workdir=args.workdir, results=results)
        print(report)
        return
    if args.settings_only:
        if args.all_sizes:
            raise ValueError("--settings-only запускается для одного --size.")
        rows = int(args.rows if args.rows is not None else DEFAULT_SIZES[args.size])
        results = run_settings_benchmark(
            workdir=args.workdir,
            size=args.size,
            rows=rows,
            merge_size=args.settings_merge_size,
            include_merge=not args.settings_skip_merge,
        )
        report = write_settings_outputs(
            workdir=args.workdir,
            size=args.size,
            rows=rows,
            merge_size=args.settings_merge_size,
            results=results,
        )
        print(f"\n## DuckDB settings: {args.size}, {rows:,} rows")
        print(report)
        return
    if args.merge_only:
        if args.rows is not None:
            raise ValueError("--rows не применяется к --merge-only; размеры merge benchmark фиксированы.")
        sizes = merge_sizes_for_args(size=args.size, all_sizes=args.all_sizes, include_large_merge=args.include_large_merge)
        results = [run_merge_benchmark(workdir=args.workdir, size=size) for size in sizes]
        args.workdir.mkdir(parents=True, exist_ok=True)
        payload = {
            "benchmark": "merge_csv",
            "sizes": sizes,
            "results": [
                {
                    **item.__dict__,
                    "speedup": item.speedup,
                }
                for item in results
            ],
        }
        (args.workdir / "merge_benchmark.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        report = merge_markdown_table(results)
        (args.workdir / "merge_benchmark.md").write_text(report + "\n", encoding="utf-8")
        print(report)
        return
    sizes = [size for size in BENCH_SIZE_ORDER if size in DEFAULT_SIZES] if args.all_sizes else [args.size]
    for size in sizes:
        rows = int(DEFAULT_SIZES[size] if args.all_sizes else args.rows if args.rows is not None else DEFAULT_SIZES[size])
        results, quality = run_benchmark(
            workdir=args.workdir,
            size=size,
            rows=rows,
            threads=args.threads,
            memory_limit=args.memory_limit,
            skip_excel=args.skip_excel,
        )
        report = write_outputs(workdir=args.workdir, size=size, rows=rows, results=results, quality=quality)
        print(f"\n## {size}: {rows:,} rows")
        print(report)
        print("\nQuality:")
        print(json.dumps(quality, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
